"""
OncoConnect Co-Creation Hub v4.0
Erasmus+ KA210 — AI-Driven Proposal Governance Platform
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json, io
from datetime import datetime, timedelta, date, time as dt_time
from io import BytesIO

# ═══════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════
st.set_page_config(
    page_title="OncoConnect Co-Creation Hub",
    page_icon="🧬", layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════
# SUPABASE
# ═══════════════════════════════════════
SUPABASE_OK = False
try:
    from supabase import create_client, Client

    @st.cache_resource
    def get_supabase() -> Client:
        return create_client(
            st.secrets["supabase"]["url"],
            st.secrets["supabase"]["key"]
        )

    _client = get_supabase()
    SUPABASE_OK = True
except Exception:
    SUPABASE_OK = False


def sb():
    return get_supabase() if SUPABASE_OK else None

# ═══════════════════════════════════════
# AI ENGINE
# ═══════════════════════════════════════
AI_ENABLED = False
ai_client = None
ai_model = "gpt-4o-mini"
try:
    import openai
    ork = st.secrets.get("openrouter", {}).get("api_key", "")
    oak = st.secrets.get("openai", {}).get("api_key", "")
    if ork:
        ai_client = openai.OpenAI(
            api_key=ork,
            base_url="https://openrouter.ai/api/v1"
        )
        ai_model = st.secrets.get("openrouter", {}).get(
            "model", "openai/gpt-4o-mini"
        )
        AI_ENABLED = True
    elif oak:
        ai_client = openai.OpenAI(api_key=oak)
        AI_ENABLED = True
except Exception:
    pass

# ═══════════════════════════════════════
# EXPORT LIBS
# ═══════════════════════════════════════
EXPORT_OK = False
try:
    from fpdf import FPDF
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXPORT_OK = True
except ImportError:
    pass

# ═══════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════
SUBMISSION_DEADLINE = datetime(2027, 5, 15, 17, 0, 0)
PREPARATION_START = datetime(2025, 6, 5)
PROJECT_START = datetime(2025, 9, 1)
TOTAL_BUDGET = 60000
BUCKET = "oncoconnect-files"

PARTNER_MAP = {
    "Turkey": "Kanser Savaşçıları Derneği",
    "Poland": "Fundacja Onkologiczna Rakiety",
    "Spain": "Universitat de Barcelona",
}
FLAGS = {"Turkey": "🇹🇷", "Poland": "🇵🇱", "Spain": "🇪🇸"}
ROLE_BADGES = {
    "Admin": "🛡️ Admin",
    "Partner": "🤝 Partner",
    "Patient": "💚 Patient",
}

PROPOSAL_SECTIONS = [
    "Project Summary", "Problem Analysis", "Objectives",
    "Methodology", "Work Packages", "Partnership",
    "Impact", "Evaluation", "Budget",
    "Dissemination", "Ethics / GDPR", "Sustainability",
]

PROPOSAL_SECTION_KEYS = [
    ("project_summary", "Project Summary"),
    ("problem_analysis", "Problem Analysis"),
    ("objectives", "Objectives"),
    ("methodology", "Methodology"),
    ("work_packages", "Work Packages"),
    ("partnership", "Partnership"),
    ("impact", "Impact"),
    ("evaluation", "Evaluation"),
    ("budget", "Budget"),
    ("dissemination", "Dissemination"),
    ("ethics_gdpr", "Ethics / GDPR"),
    ("sustainability", "Sustainability"),
]

USERS_DB = {
    "admin": {
        "password": "admin123", "name": "Project Admin",
        "role": "Admin", "country": "All",
        "org": "OncoConnect", "can_read_patient_fb": True,
    },
    "turkey": {
        "password": "tr2025", "name": "KSD Coordinator",
        "role": "Partner", "country": "Turkey",
        "org": "Kanser Savaşçıları Derneği",
        "can_read_patient_fb": False,
    },
    "poland": {
        "password": "pl2025", "name": "Rakiety Team",
        "role": "Partner", "country": "Poland",
        "org": "Fundacja Onkologiczna Rakiety",
        "can_read_patient_fb": False,
    },
    "spain": {
        "password": "es2025", "name": "UB Research Team",
        "role": "Partner", "country": "Spain",
        "org": "Universitat de Barcelona",
        "can_read_patient_fb": False,
    },
    "patient": {
        "password": "patient123", "name": "Patient Participant",
        "role": "Patient", "country": "N/A",
        "org": "N/A", "can_read_patient_fb": False,
    },
}

WP_COUNTRY_MAP = {
    "Turkey": {"lead": ["WP1", "WP4"], "support": ["WP2", "WP3", "WP5"]},
    "Poland": {"lead": ["WP2"], "support": ["WP3", "WP4", "WP5"]},
    "Spain": {"lead": ["WP3", "WP5"], "support": ["WP2", "WP4"]},
}

PAGE_PERMISSIONS = {
    "Dashboard": {"Admin": "full", "Partner": "read", "Patient": "read"},
    "Work Packages": {"Admin": "full", "Partner": "filtered", "Patient": "none"},
    "Gantt Chart": {"Admin": "full", "Partner": "filtered", "Patient": "none"},
    "Partners": {"Admin": "full", "Partner": "read", "Patient": "read"},
    "Partner Feedback": {"Admin": "full", "Partner": "write", "Patient": "none"},
    "Patient Feedback": {"Admin": "full", "Partner": "none", "Patient": "write"},
    "Approval Status": {"Admin": "full", "Partner": "own", "Patient": "none"},
    "Announcements": {"Admin": "full", "Partner": "write", "Patient": "read"},
    "Documents": {"Admin": "full", "Partner": "upload", "Patient": "read"},
    "Meetings": {"Admin": "full", "Partner": "read", "Patient": "read"},
    "🧠 AI Center": {"Admin": "full", "Partner": "none", "Patient": "none"},
    "Admin Panel": {"Admin": "full", "Partner": "none", "Patient": "none"},
    "User Management": {"Admin": "full", "Partner": "none", "Patient": "none"},
}


# ═══════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════
def get_user_wps(country):
    if country == "All":
        return ["WP1", "WP2", "WP3", "WP4", "WP5"]
    m = WP_COUNTRY_MAP.get(country, {})
    return m.get("lead", []) + m.get("support", [])


def get_wp_role(country, wp):
    m = WP_COUNTRY_MAP.get(country, {})
    if wp in m.get("lead", []):
        return "🟢 Lead"
    if wp in m.get("support", []):
        return "🔵 Support"
    return "⚪ —"


def check_access(page, role):
    return PAGE_PERMISSIONS.get(page, {}).get(role, "none") != "none"


def get_permission(page, role):
    return PAGE_PERMISSIONS.get(page, {}).get(role, "none")


# ═══════════════════════════════════════
# CSS
# ═══════════════════════════════════════
def inject_pro_css():
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    .stApp { font-family: 'Inter', sans-serif; }

    .pro-header {
        background: linear-gradient(135deg, #1B3A5C 0%, #2d5a8e 50%, #2ABFBF 100%);
        padding: 2rem 2.5rem; border-radius: 20px; margin-bottom: 2rem;
        color: white; position: relative; overflow: hidden;
        box-shadow: 0 10px 40px rgba(27,58,92,0.3);
    }
    .pro-header h1 { margin:0; font-size:2rem; font-weight:800; }
    .pro-header p { margin:0.3rem 0 0; opacity:0.85; font-size:0.95rem; }

    .metric-card {
        background: linear-gradient(145deg, #ffffff, #f8f9fc);
        border: 1px solid #e8ecf1; border-radius: 16px;
        padding: 1.5rem; text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.05);
        border-top: 4px solid #2ABFBF;
    }
    .metric-card .value { font-size:2.2rem; font-weight:800; color:#1B3A5C; }
    .metric-card .label { font-size:0.8rem; color:#8896a6; margin-top:0.3rem; font-weight:500; text-transform:uppercase; letter-spacing:0.5px; }

    .meeting-card {
        background: white; border: 1px solid #e8ecf1;
        border-radius: 16px; padding: 1.5rem; margin-bottom: 1rem;
        border-left: 5px solid #2ABFBF; box-shadow: 0 2px 10px rgba(0,0,0,0.04);
    }
    .meeting-card.past { border-left-color: #94a3b8; opacity: 0.7; }
    .meeting-card.today { border-left-color: #10B981; background: linear-gradient(135deg, #f0fdf4, #ffffff); }
    .meeting-date { font-size:0.8rem; color:#64748b; font-weight:600; text-transform:uppercase; letter-spacing:1px; }
    .meeting-title { font-size:1.2rem; font-weight:700; color:#1B3A5C; margin:0.3rem 0; }
    .meeting-time { font-size:1rem; color:#2ABFBF; font-weight:600; }

    .zoom-btn {
        display:inline-block; background:linear-gradient(135deg,#2D8CFF,#2681F2);
        color:white !important; padding:8px 20px; border-radius:10px;
        font-weight:600; font-size:0.85rem; text-decoration:none;
        box-shadow:0 3px 10px rgba(45,140,255,0.3);
    }

    .badge { display:inline-block; padding:4px 12px; border-radius:20px; font-size:0.75rem; font-weight:600; }
    .badge-success { background:#dcfce7; color:#16a34a; }
    .badge-warning { background:#fef9c3; color:#ca8a04; }
    .badge-danger { background:#fee2e2; color:#dc2626; }
    .badge-info { background:#dbeafe; color:#2563eb; }

    .fb-card {
        background: white; border: 1px solid #e8ecf1;
        border-radius: 12px; padding: 1.2rem; margin: 0.5rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.03);
    }
    .ann-card {
        border-left: 4px solid; padding: 1rem 1.2rem; margin-bottom: 0.7rem;
        background: white; border-radius: 0 12px 12px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    .access-denied {
        background: linear-gradient(135deg, #fef2f2, #fff1f2);
        border: 1px solid #fecaca; border-radius: 16px;
        padding: 3rem; text-align: center; margin: 2rem 0;
    }

    div[data-testid="stSidebar"] { background: linear-gradient(180deg, #0a1628 0%, #1B3A5C 100%); }
    div[data-testid="stSidebar"] .stMarkdown { color: #cbd5e1; }
    .stButton > button { border-radius: 10px; font-weight: 600; }
    </style>""", unsafe_allow_html=True)


# ═══════════════════════════════════════
# AUTH
# ═══════════════════════════════════════
def init_session():
    defaults = {
        "authenticated": False, "username": None,
        "user_name": None, "user_role": None,
        "user_country": None, "user_org": None,
        "can_read_patient_fb": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def render_login():
    if st.session_state.get("authenticated"):
        return True
    inject_pro_css()
    st.markdown(
        "<div style='text-align:center;padding:3rem 0 1rem;'>"
        "<h1 style='font-size:3rem;'>🧬 OncoConnect</h1>"
        "<h3 style='color:#64748b;font-weight:400;'>Co-Creation Hub</h3>"
        "<p style='color:#94a3b8;'>Erasmus+ KA210</p></div>",
        unsafe_allow_html=True,
    )
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("login"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            if st.form_submit_button("Login", use_container_width=True, type="primary"):
                ud = None
                # Try Supabase first
                if SUPABASE_OK:
                    try:
                        res = (
                            sb()
                            .table("app_users")
                            .select("*")
                            .eq("username", username.strip().lower())
                            .eq("is_active", True)
                            .execute()
                        )
                        if res.data:
                            u = res.data[0]
                            if u["password_hash"] == password:
                                ud = {
                                    "name": u["display_name"],
                                    "role": u["role"],
                                    "country": u["country"],
                                    "org": u.get("organisation", "N/A"),
                                    "can_read_patient_fb": u.get(
                                        "can_read_patient_fb", False
                                    ),
                                }
                    except Exception:
                        pass
                # Fallback
                if not ud:
                    u = USERS_DB.get(username)
                    if u and u["password"] == password:
                        ud = {
                            "name": u["name"],
                            "role": u["role"],
                            "country": u["country"],
                            "org": u["org"],
                            "can_read_patient_fb": u.get(
                                "can_read_patient_fb", False
                            ),
                        }
                if ud:
                    st.session_state.update(
                        authenticated=True,
                        username=username.strip().lower(),
                        user_name=ud["name"],
                        user_role=ud["role"],
                        user_country=ud["country"],
                        user_org=ud["org"],
                        can_read_patient_fb=ud["can_read_patient_fb"],
                    )
                    st.rerun()
                else:
                    st.error("Invalid credentials")
        with st.expander("Demo Credentials"):
            st.markdown(
                "| User | Pass | Role |\n|---|---|---|\n"
                "| admin | admin123 | Admin |\n"
                "| turkey | tr2025 | Partner TR |\n"
                "| poland | pl2025 | Partner PL |\n"
                "| spain | es2025 | Partner ES |\n"
                "| patient | patient123 | Patient |"
            )
    return False


def logout():
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()


def get_role():
    return st.session_state.get("user_role", "Patient")


def get_country():
    return st.session_state.get("user_country", "N/A")


def get_name():
    return st.session_state.get("user_name", "User")


def get_org():
    return st.session_state.get("user_org", "N/A")


# ═══════════════════════════════════════
# STORAGE
# ═══════════════════════════════════════
def upload_to_storage(fb, path, ct="application/octet-stream"):
    if not SUPABASE_OK:
        return False
    try:
        sb().storage.from_(BUCKET).upload(
            path=path,
            file=fb,
            file_options={"content-type": ct, "upsert": "true"},
        )
        return True
    except Exception as e:
        st.error(f"Upload error: {e}")
        return False


def download_from_storage(path):
    if not SUPABASE_OK:
        return None
    try:
        return sb().storage.from_(BUCKET).download(path)
    except Exception:
        return None


def delete_from_storage(path):
    if not SUPABASE_OK:
        return False
    try:
        sb().storage.from_(BUCKET).remove([path])
        return True
    except Exception:
        return False


def save_document_metadata(fn, ft, fs, cat, desc, ub, co, sp, ver=1):
    if not SUPABASE_OK:
        return
    try:
        sb().table("documents").insert(
            {
                "file_name": fn, "file_type": ft, "file_size": fs,
                "category": cat, "description": desc, "uploaded_by": ub,
                "country": co, "storage_path": sp, "version": ver,
                "is_active": True,
            }
        ).execute()
    except Exception as e:
        st.error(f"Error: {e}")


# ═══════════════════════════════════════
# DATABASE FUNCTIONS
# ═══════════════════════════════════════
def db_get_approvals():
    d = {"Turkey": False, "Poland": False, "Spain": False}
    if not SUPABASE_OK:
        return st.session_state.get("local_approvals", d)
    try:
        r = sb().table("approvals").select("country, approved").execute()
        res = {x["country"]: x.get("approved", False) for x in (r.data or [])}
        for c in d:
            res.setdefault(c, False)
        return res
    except Exception:
        return d


def db_set_approval(cn, approved, by, role):
    now = datetime.utcnow().isoformat()
    if not SUPABASE_OK:
        st.session_state.setdefault(
            "local_approvals",
            {"Turkey": False, "Poland": False, "Spain": False},
        )[cn] = approved
        return
    try:
        sb().table("approvals").update(
            {
                "approved": approved,
                "status": "Approved" if approved else "Pending",
                "approved_by": by if approved else None,
                "approved_at": now if approved else None,
                "updated_at": now,
            }
        ).eq("country", cn).execute()
        sb().table("approval_log").insert(
            {
                "action": "approved" if approved else "revoked",
                "country": cn, "performed_by": by, "role": role,
            }
        ).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_reset_all_approvals(by):
    for c in ["Turkey", "Poland", "Spain"]:
        db_set_approval(c, False, by, "Admin")


def db_get_approval_log():
    if not SUPABASE_OK:
        return []
    try:
        return (
            sb()
            .table("approval_log")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_get_partner_feedback():
    if not SUPABASE_OK:
        return st.session_state.get("local_feedback", [])
    try:
        return (
            sb()
            .table("partner_feedback")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_add_partner_feedback(pc, org, sec, fb, pri, by):
    if not SUPABASE_OK:
        st.session_state.setdefault("local_feedback", []).append(
            {
                "id": len(st.session_state.get("local_feedback", [])) + 1,
                "partner_country": pc, "organisation": org,
                "section": sec, "feedback": fb, "content": fb,
                "priority": pri, "status": "Open",
                "submitted_by": by, "country": pc,
                "created_at": datetime.now().isoformat(),
            }
        )
        return
    try:
        sb().table("partner_feedback").insert(
            {
                "partner_country": pc, "organisation": org,
                "section": sec, "feedback": fb, "content": fb,
                "priority": pri, "status": "Open",
                "submitted_by": by, "country": pc,
            }
        ).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_update_feedback_status(fid, new_status, resp=None):
    if not SUPABASE_OK:
        return
    try:
        d = {"status": new_status}
        if resp:
            d["response"] = resp
        sb().table("partner_feedback").update(d).eq("id", fid).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_get_patient_feedback():
    if not SUPABASE_OK:
        return st.session_state.get("local_patient_fb", [])
    try:
        return (
            sb()
            .table("patient_feedback")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_add_patient_feedback(data):
    if not SUPABASE_OK:
        data["id"] = len(st.session_state.get("local_patient_fb", [])) + 1
        data["created_at"] = datetime.now().isoformat()
        st.session_state.setdefault("local_patient_fb", []).append(data)
        return
    try:
        sb().table("patient_feedback").insert(data).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_get_announcements():
    if not SUPABASE_OK:
        return st.session_state.get("local_ann", [])
    try:
        return (
            sb()
            .table("announcements")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_add_announcement(title, content, author, priority):
    if not SUPABASE_OK:
        st.session_state.setdefault("local_ann", []).append(
            {
                "id": len(st.session_state.get("local_ann", [])) + 1,
                "title": title, "content": content,
                "author": author, "priority": priority,
                "created_at": datetime.now().isoformat(),
            }
        )
        return
    try:
        sb().table("announcements").insert(
            {"title": title, "content": content, "author": author, "priority": priority}
        ).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_log_improvement(fid, sec, old, new, reason, action, by):
    if not SUPABASE_OK:
        st.session_state.setdefault("local_imp", []).append(
            {
                "id": len(st.session_state.get("local_imp", [])) + 1,
                "feedback_id": fid, "section": sec,
                "original_text": old, "updated_text": new,
                "ai_reasoning": reason, "action": action,
                "created_by": by,
                "created_at": datetime.now().isoformat(),
            }
        )
        return
    try:
        sb().table("improvement_log").insert(
            {
                "feedback_id": fid, "section": sec,
                "original_text": old, "updated_text": new,
                "ai_reasoning": reason, "action": action,
                "created_by": by,
            }
        ).execute()
    except Exception:
        pass


def db_get_improvement_log():
    if not SUPABASE_OK:
        return st.session_state.get("local_imp", [])
    try:
        return (
            sb()
            .table("improvement_log")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_log_ai_decision(fid, dec, conf, reason, target):
    if not SUPABASE_OK:
        st.session_state.setdefault("local_ai", []).append(
            {
                "id": len(st.session_state.get("local_ai", [])) + 1,
                "feedback_id": fid, "decision": dec,
                "confidence": conf, "reasoning": reason,
                "target_section": target,
                "created_at": datetime.now().isoformat(),
            }
        )
        return
    try:
        sb().table("ai_decisions").insert(
            {
                "feedback_id": fid, "decision": dec,
                "confidence": conf, "reasoning": reason,
                "target_section": target,
            }
        ).execute()
    except Exception:
        pass


def db_get_ai_decisions():
    if not SUPABASE_OK:
        return st.session_state.get("local_ai", [])
    try:
        return (
            sb()
            .table("ai_decisions")
            .select("*")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_get_proposal_sections():
    if not SUPABASE_OK:
        return st.session_state.get("local_sec", [])
    try:
        return (
            sb()
            .table("proposal_sections")
            .select("*")
            .eq("is_active", True)
            .order("section_order")
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_get_section_by_key(key):
    if not SUPABASE_OK:
        return next(
            (s for s in st.session_state.get("local_sec", []) if s.get("section_key") == key),
            None,
        )
    try:
        r = (
            sb()
            .table("proposal_sections")
            .select("*")
            .eq("section_key", key)
            .eq("is_active", True)
            .execute()
        )
        return r.data[0] if r.data else None
    except Exception:
        return None


def db_update_section_content(sk, content, by, fid=None):
    if not SUPABASE_OK:
        return 1
    try:
        cur = db_get_section_by_key(sk)
        nv = (cur.get("version", 1) + 1) if cur else 1
        sb().table("proposal_sections").update(
            {
                "content": content, "version": nv,
                "last_updated_by": by, "last_feedback_id": fid,
                "updated_at": datetime.utcnow().isoformat(),
            }
        ).eq("section_key", sk).eq("is_active", True).execute()
        return nv
    except Exception as e:
        st.error(f"Error: {e}")
        return None


def _find_section_for_feedback(sec):
    sk = sec.lower().replace(" ", "_").replace("/", "_").replace("__", "_").strip("_")
    sd = db_get_section_by_key(sk)
    if sd:
        return sd
    for s in db_get_proposal_sections():
        if s["section_title"].lower() == sec.lower() or sec.lower() in s["section_title"].lower():
            return s
    return None


# ─── MEETINGS DB ───
def db_get_meetings():
    if not SUPABASE_OK:
        return st.session_state.get("local_meetings", [])
    try:
        return (
            sb()
            .table("meetings")
            .select("*")
            .order("meeting_date", desc=False)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def db_add_meeting(data):
    if not SUPABASE_OK:
        data["id"] = len(st.session_state.get("local_meetings", [])) + 1
        data["created_at"] = datetime.now().isoformat()
        st.session_state.setdefault("local_meetings", []).append(data)
        return
    try:
        sb().table("meetings").insert(data).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_update_meeting(mid, data):
    if not SUPABASE_OK:
        return
    try:
        sb().table("meetings").update(data).eq("id", mid).execute()
    except Exception as e:
        st.error(f"Error: {e}")


def db_delete_meeting(mid):
    if not SUPABASE_OK:
        return
    try:
        sb().table("meetings").delete().eq("id", mid).execute()
    except Exception as e:
        st.error(f"Error: {e}")


# ═══════════════════════════════════════
# AI FUNCTIONS
# ═══════════════════════════════════════
def ai_analyze_feedback_v2(ft, sk, sc):
    if not AI_ENABLED or not ai_client:
        return {
            "decision": "manual_review", "confidence": 0,
            "reasoning": "AI not configured.",
            "target_section": sk, "suggested_action": "review",
            "suggested_text": "", "priority": "medium",
            "affected_wp": "", "erasmus_criteria": {},
        }
    try:
        r = ai_client.chat.completions.create(
            model=ai_model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are an Erasmus+ KA210 proposal analyst for OncoConnect. "
                        'Return JSON: {"decision":"integrate|revise|route|archive|reject|ethical_risk",'
                        '"confidence":0.0-1.0,"reasoning":"...","target_section":"...",'
                        '"suggested_action":"...","suggested_text":"...","priority":"critical|high|medium|low",'
                        '"affected_wp":"WP1-WP5","erasmus_criteria":{}}'
                    ),
                },
                {"role": "user", "content": f"SECTION: {sk}\nCONTENT:\n{sc or '[Empty]'}\nFEEDBACK:\n{ft}"},
            ],
            response_format={"type": "json_object"},
            temperature=0.3,
            max_tokens=1500,
        )
        return json.loads(r.choices[0].message.content)
    except Exception as e:
        return {
            "decision": "manual_review", "confidence": 0,
            "reasoning": f"AI error: {e}", "target_section": sk,
            "suggested_action": "review", "suggested_text": "",
            "priority": "medium", "affected_wp": "",
            "erasmus_criteria": {},
        }


def ai_generate_section_revision(sk, sc, fb):
    if not AI_ENABLED or not ai_client:
        return "AI not configured."
    try:
        r = ai_client.chat.completions.create(
            model=ai_model,
            messages=[
                {"role": "system", "content": "Expert Erasmus+ KA210 proposal writer. Output ONLY revised text."},
                {"role": "user", "content": f"SECTION: {sk}\nCURRENT:\n{sc or '[Empty]'}\nFEEDBACK:\n{fb}\nWrite improved version:"},
            ],
            temperature=0.4,
            max_tokens=2000,
        )
        return r.choices[0].message.content
    except Exception as e:
        return f"AI error: {e}"


def ai_generate_summary(fbl):
    if not AI_ENABLED or not ai_client or not fbl:
        return "AI not available."
    try:
        ft = "\n".join(
            [f"- [{f.get('section','')}] {f.get('feedback', f.get('content',''))}" for f in fbl[:20]]
        )
        r = ai_client.chat.completions.create(
            model=ai_model,
            messages=[
                {"role": "system", "content": "Summarize key themes from Erasmus+ KA210 feedback."},
                {"role": "user", "content": ft},
            ],
            temperature=0.3,
            max_tokens=800,
        )
        return r.choices[0].message.content
    except Exception as e:
        return f"AI error: {e}"


# ═══════════════════════════════════════
# PROPOSAL PARSER
# ═══════════════════════════════════════
def parse_proposal_md(md):
    mapping = {
        "project summary": "project_summary",
        "problem analysis": "problem_analysis",
        "needs analysis": "problem_analysis",
        "objectives": "objectives",
        "methodology": "methodology",
        "work packages": "work_packages",
        "partnership": "partnership",
        "impact": "impact",
        "evaluation": "evaluation",
        "budget": "budget",
        "dissemination": "dissemination",
        "ethics": "ethics_gdpr",
        "gdpr": "ethics_gdpr",
        "sustainability": "sustainability",
    }
    result, ck, cl = {}, None, []
    for line in md.split("\n"):
        s = line.strip()
        if s.startswith("## ") or s.startswith("# "):
            if ck and cl:
                result[ck] = "\n".join(cl).strip()
            t = s.lstrip("# ").strip().lower()
            ck = None
            for kw, sk in mapping.items():
                if kw in t:
                    ck = sk
                    break
            cl = []
        elif ck:
            cl.append(line)
    if ck and cl:
        result[ck] = "\n".join(cl).strip()
    return result


# ═══════════════════════════════════════
# EXPORT ENGINE
# ═══════════════════════════════════════
def generate_feedback_excel():
    if not EXPORT_OK:
        return None
    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfl = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    bd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wa = Alignment(wrap_text=True, vertical="top")
    pfb = db_get_partner_feedback()
    patfb = db_get_patient_feedback()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "OncoConnect — Feedback Export Report"
    ws["A1"].font = Font(bold=True, size=16, color="1B3A5C")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    stats = [
        ("Total Partner Feedback", len(pfb)),
        ("Total Patient Feedback", len(patfb)),
        ("Open", len([f for f in pfb if f.get("status") == "Open"])),
        ("Accepted", len([f for f in pfb if f.get("status") == "Accepted"])),
        ("Rejected", len([f for f in pfb if f.get("status") == "Rejected"])),
    ]
    for i, (l, v) in enumerate(stats, 5):
        ws[f"A{i}"] = l
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = v
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

    # All partner feedback sheet
    ws2 = wb.create_sheet("All Partner Feedback")
    headers = ["ID", "Date", "Country", "Organisation", "Section", "Priority", "Status", "Feedback", "Response"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfl
        cell.border = bd
    for i, fb in enumerate(pfb, 2):
        vals = [
            fb.get("id", ""),
            str(fb.get("created_at", ""))[:10],
            fb.get("partner_country", fb.get("country", "")),
            fb.get("organisation", ""),
            fb.get("section", ""),
            fb.get("priority", ""),
            fb.get("status", ""),
            fb.get("feedback", fb.get("content", "")),
            fb.get("response", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=str(v) if v else "")
            cell.border = bd
            cell.alignment = wa
    for i, w in enumerate([8, 12, 12, 30, 20, 10, 12, 60, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Patient feedback sheet
    if patfb:
        wsp = wb.create_sheet("Patient Feedback")
        ph = ["ID", "Date", "Country", "Age", "Cancer", "Support Need", "Digital", "Matching", "Privacy"]
        for c, h in enumerate(ph, 1):
            cell = wsp.cell(row=1, column=c, value=h)
            cell.font = hf
            cell.fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
            cell.border = bd
        for i, p in enumerate(patfb, 2):
            vals = [
                p.get("id", ""), str(p.get("created_at", ""))[:10],
                p.get("country", ""), p.get("age_group", ""),
                p.get("cancer_type", ""), p.get("support_need", ""),
                p.get("digital_literacy", ""), p.get("matching_preference", ""),
                p.get("privacy_expectation", ""),
            ]
            for c, v in enumerate(vals, 1):
                cell = wsp.cell(row=i, column=c, value=str(v) if v else "")
                cell.border = bd

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_feedback_pdf():
    if not EXPORT_OK:
        return None

    class P(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(27, 58, 92)
            self.cell(0, 8, "OncoConnect - Feedback Report", align="L")
            self.cell(0, 8, datetime.now().strftime("%d %B %Y"), align="R", new_x="LMARGIN", new_y="NEXT")
            self.set_draw_color(42, 191, 191)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f"OncoConnect | Page {self.page_no()}/{{nb}}", align="C")

        def stitle(self, t):
            self.set_font("Helvetica", "B", 14)
            self.set_text_color(27, 58, 92)
            self.set_fill_color(240, 242, 246)
            self.cell(0, 10, t, fill=True, new_x="LMARGIN", new_y="NEXT")
            self.ln(3)

        def body(self, t):
            self.set_font("Helvetica", "", 9)
            self.set_text_color(51, 51, 51)
            safe = t.encode("latin-1", "replace").decode("latin-1") if t else ""
            self.multi_cell(0, 5, safe)
            self.ln(2)

    pdf = P()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pfb = db_get_partner_feedback()
    patfb = db_get_patient_feedback()
    ap = db_get_approvals()

    # Cover page
    pdf.add_page()
    pdf.ln(30)
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(27, 58, 92)
    pdf.cell(0, 15, "OncoConnect", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(42, 191, 191)
    pdf.cell(0, 10, "Feedback & AI Analysis Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(128, 128, 128)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(51, 51, 51)
    an = sum(1 for v in ap.values() if v)
    for l in [f"Partner Feedback: {len(pfb)}", f"Patient Feedback: {len(patfb)}", f"Approvals: {an}/3"]:
        pdf.cell(0, 7, l, align="C", new_x="LMARGIN", new_y="NEXT")

    # Sections
    secs = sorted(set(f.get("section", "General") for f in pfb))
    for sn in secs:
        pdf.add_page()
        pdf.stitle(f"Section: {sn}")
        sfbs = [f for f in pfb if f.get("section") == sn]
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(27, 58, 92)
        pdf.cell(0, 8, f"Feedback ({len(sfbs)} items):", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        for fb in sfbs:
            fc = fb.get("partner_country", fb.get("country", ""))
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 6, f"#{fb.get('id','')} | {fc} | {fb.get('priority','')} | {fb.get('status','')}", new_x="LMARGIN", new_y="NEXT")
            pdf.body(fb.get("feedback", fb.get("content", "")))
            if pdf.get_y() > 260:
                pdf.add_page()

    return bytes(pdf.output())


# ═══════════════════════════════════════
# DATA LOADERS & UI HELPERS
# ═══════════════════════════════════════
@st.cache_data
def load_csv(p):
    try:
        return pd.read_csv(p)
    except Exception:
        return pd.DataFrame()


def load_static():
    return load_csv("data/work_packages.csv"), load_csv("data/partners.csv")


def render_countdown():
    now = datetime.now()
    pt = (SUBMISSION_DEADLINE - PREPARATION_START).days
    pe = (now - PREPARATION_START).days
    pp = max(0, min(1, pe / max(pt, 1)))
    rem = SUBMISSION_DEADLINE - now
    if rem.total_seconds() <= 0:
        st.error("DEADLINE PASSED!")
        return
    d = rem.days
    h, r2 = divmod(rem.seconds, 3600)
    m, _ = divmod(r2, 60)
    sc = "#17a2b8" if d > 365 else "#28a745" if d > 180 else "#ffc107" if d > 60 else "#dc3545"
    pct = int(pp * 100)
    deg = int(pp * 360)
    pc = "#a855f7" if pp < 0.5 else "#f59e0b" if pp < 0.8 else "#ef4444"
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(
            f"<div style='background:linear-gradient(135deg,#1a1a2e,#2d1b4e);border-radius:16px;"
            f"padding:1.5rem;text-align:center;color:white;min-height:250px;'>"
            f"<p style='margin:0;font-size:.75rem;opacity:.7;letter-spacing:2px;'>PREPARATION PHASE</p>"
            f"<div style='margin:1rem auto;width:120px;height:120px;border-radius:50%;"
            f"background:conic-gradient({pc} {deg}deg,#333 0deg);display:flex;align-items:center;justify-content:center;'>"
            f"<div style='width:100px;height:100px;border-radius:50%;background:#1a1a2e;"
            f"display:flex;align-items:center;justify-content:center;flex-direction:column;'>"
            f"<span style='font-size:1.8rem;font-weight:bold;color:{pc};'>{pct}%</span>"
            f"<span style='font-size:.65rem;opacity:.6;'>COMPLETE</span></div></div>"
            f"<p style='margin:0;font-size:.8rem;opacity:.6;'>Elapsed: {pe} days</p></div>",
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"<div style='background:linear-gradient(135deg,#0f2027,#203a43,#2c5364);border-radius:16px;"
            f"padding:1.5rem;text-align:center;color:white;min-height:250px;'>"
            f"<p style='margin:0;font-size:.75rem;opacity:.7;letter-spacing:2px;'>SUBMISSION DEADLINE</p>"
            f"<p style='margin:.3rem 0 0;font-size:.85rem;opacity:.8;'>{SUBMISSION_DEADLINE.strftime('%d %B %Y')}</p>"
            f"<div style='display:flex;justify-content:center;gap:1.5rem;margin:1.2rem 0;'>"
            f"<div><span style='font-size:2.8rem;font-weight:bold;color:{sc};'>{d}</span><br>"
            f"<span style='font-size:.75rem;opacity:.6;'>DAYS</span></div>"
            f"<div><span style='font-size:2.8rem;font-weight:bold;color:{sc};'>{h:02d}</span><br>"
            f"<span style='font-size:.75rem;opacity:.6;'>HOURS</span></div>"
            f"<div><span style='font-size:2.8rem;font-weight:bold;color:{sc};'>{m:02d}</span><br>"
            f"<span style='font-size:.75rem;opacity:.6;'>MIN</span></div></div></div>",
            unsafe_allow_html=True,
        )
    st.progress(max(0, min(1, 1 - d / max(pt, 1))))


def ann_card(row):
    p = row.get("priority", "Low")
    border = {"High": "#dc3545", "Medium": "#ffc107"}.get(p, "#28a745")
    icon = {"High": "🔴", "Medium": "🟡"}.get(p, "🟢")
    st.markdown(
        f"<div class='ann-card' style='border-left-color:{border};'>"
        f"<strong>{icon} {row.get('title','')}</strong>"
        f"<span style='float:right;color:#666;font-size:.85rem;'>{str(row.get('created_at',''))[:10]}</span><br>"
        f"<span style='color:#444;'>{row.get('content','')}</span><br>"
        f"<span style='font-size:.8rem;color:#999;'>By: {row.get('author','')}</span></div>",
        unsafe_allow_html=True,
    )


def show_access_denied():
    st.markdown(
        "<div class='access-denied'><h2>🔒 Access Denied</h2>"
        "<p>You don't have permission.</p></div>",
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════
# PAGE: DASHBOARD
# ═══════════════════════════════════════
def page_dashboard(wp_df, pf):
    r = get_role()
    c = get_country()
    st.markdown(
        "<div class='pro-header'><h1>🧬 OncoConnect Co-Creation Hub</h1>"
        "<p>Erasmus+ KA210 — AI-Driven Proposal Governance Platform</p></div>",
        unsafe_allow_html=True,
    )
    render_countdown()
    ap = db_get_approvals()
    fbc = len(db_get_partner_feedback())
    aic = len(db_get_ai_decisions())
    an = sum(1 for v in ap.values() if v)
    try:
        dc = len(sb().table("documents").select("id").eq("is_active", True).execute().data or []) if SUPABASE_OK else 0
    except Exception:
        dc = 0

    meetings = db_get_meetings()
    today = date.today()
    upcoming = [m for m in meetings if str(m.get("meeting_date", "")) >= str(today)]

    cols = st.columns(6)
    metrics = [
        ("Work Packages", len(wp_df) if len(wp_df) > 0 else 5),
        ("Partners", 3),
        ("Feedback", fbc),
        ("AI Decisions", aic),
        ("Approvals", f"{an}/3"),
        ("Meetings", len(upcoming)),
    ]
    for col, (label, val) in zip(cols, metrics):
        with col:
            st.markdown(
                f"<div class='metric-card'><div class='value'>{val}</div>"
                f"<div class='label'>{label}</div></div>",
                unsafe_allow_html=True,
            )

    st.markdown("")
    sc1, sc2, sc3, sc4 = st.columns(4)
    with sc1:
        (st.success if SUPABASE_OK else st.warning)(f"🔗 DB: {'Connected' if SUPABASE_OK else 'Local'}")
    with sc2:
        (st.success if AI_ENABLED else st.warning)(f"🧠 AI: {'Active' if AI_ENABLED else 'Inactive'}")
    with sc3:
        (st.success if an == 3 else st.warning)(f"🗳️ Approvals: {an}/3")
    with sc4:
        if upcoming:
            st.info(f"📅 Next: {upcoming[0].get('title','')} ({upcoming[0].get('meeting_date','')})")
        else:
            st.info("📅 No upcoming meetings")

    if r == "Partner" and c in WP_COUNTRY_MAP:
        st.info(f"🔒 **{c}** — WPs: {', '.join(get_user_wps(c))} | Lead: {', '.join(WP_COUNTRY_MAP[c]['lead'])}")

    st.subheader("🗳️ Partner Approval")
    ac = st.columns(3)
    for col, cn in zip(ac, ["Turkey", "Poland", "Spain"]):
        with col:
            (st.success if ap.get(cn) else st.warning)(
                f"{FLAGS[cn]} {PARTNER_MAP[cn]} — {'✅' if ap.get(cn) else '⏳'}"
            )

    if len(wp_df) > 0:
        ch1, ch2 = st.columns(2)
        with ch1:
            if "budget_eur" in wp_df.columns:
                # wp_id sütunu var, wp_name'i label olarak kullan
                label_col = "wp_name" if "wp_name" in wp_df.columns else "wp_id"
                fig = px.pie(
                    wp_df, names=label_col, values="budget_eur",
                    title="Budget Distribution",
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                fig.update_traces(textposition="inside", textinfo="percent+label")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
        with ch2:
            if "status" in wp_df.columns:
                scc = wp_df["status"].value_counts().reset_index()
                scc.columns = ["status", "count"]
                fig2 = px.bar(
                    scc, x="status", y="count", color="status",
                    title="WP Status Overview",
                    color_discrete_map={
                        "Completed": "#10B981",
                        "In Progress": "#3B82F6",
                        "Not Started": "#94A3B8",
                        "Planned": "#F59E0B",
                    },
                )
                fig2.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

    # Budget summary
    if len(wp_df) > 0 and "budget_eur" in wp_df.columns:
        total_budget = wp_df["budget_eur"].sum()
        st.markdown(
            f"<div style='background:linear-gradient(135deg,#f0f9ff,#e0f2fe);border-radius:12px;"
            f"padding:1rem 1.5rem;border:1px solid #bae6fd;margin:1rem 0;'>"
            f"<strong>💰 Total Budget:</strong> €{total_budget:,.0f} / €{TOTAL_BUDGET:,.0f} "
            f"({'✅ Within limit' if total_budget <= TOTAL_BUDGET else '⚠️ Over budget'})"
            f"</div>",
            unsafe_allow_html=True,
        )

    st.subheader("📢 Latest Announcements")
    for a in db_get_announcements()[:3]:
        ann_card(a)


# ═══════════════════════════════════════
# PAGE: MEETINGS
# ═══════════════════════════════════════
def page_meetings():
    r = get_role()
    st.markdown(
        "<div class='pro-header'><h1>📅 Meetings & Zoom Sessions</h1>"
        "<p>Project coordination meetings, partner calls, and training sessions</p></div>",
        unsafe_allow_html=True,
    )

    meetings = db_get_meetings()
    today = date.today()

    tabs_list = ["📅 Upcoming", "📋 All Meetings", "📊 Calendar View"]
    if r == "Admin":
        tabs_list.insert(2, "➕ Schedule")
    tabs = st.tabs(tabs_list)

    # ─── UPCOMING ───
    with tabs[0]:
        upcoming = sorted(
            [m for m in meetings if str(m.get("meeting_date", "")) >= str(today)],
            key=lambda x: (str(x.get("meeting_date", "")), str(x.get("start_time", ""))),
        )
        if not upcoming:
            st.markdown(
                "<div style='text-align:center;padding:3rem;background:linear-gradient(135deg,#f0f9ff,#e0f2fe);"
                "border-radius:16px;border:1px solid #bae6fd;'>"
                "<h2>📭 No Upcoming Meetings</h2>"
                "<p style='color:#64748b;'>No meetings scheduled.</p></div>",
                unsafe_allow_html=True,
            )
        else:
            for m in upcoming:
                md = str(m.get("meeting_date", ""))
                is_today = md == str(today)
                card_class = "meeting-card today" if is_today else "meeting-card"
                today_badge = "<span class='badge badge-success' style='margin-left:8px;'>TODAY</span>" if is_today else ""

                zoom_link = m.get("zoom_link", "")
                zoom_btn = (
                    f"<a href='{zoom_link}' target='_blank' class='zoom-btn'>🎥 Join Zoom</a>"
                    if zoom_link
                    else "<span style='color:#94a3b8;font-size:0.85rem;'>No Zoom link</span>"
                )

                zoom_info = ""
                if m.get("zoom_id"):
                    zoom_info += f"<br><span style='font-size:0.8rem;color:#64748b;'>ID: <code>{m['zoom_id']}</code></span>"
                if m.get("zoom_passcode"):
                    zoom_info += f" <span style='font-size:0.8rem;color:#64748b;'>Pass: <code>{m['zoom_passcode']}</code></span>"

                participants = m.get("participants", '["All"]')
                if isinstance(participants, str):
                    try:
                        participants = json.loads(participants)
                    except Exception:
                        participants = [participants]
                part_str = ", ".join(participants) if participants else "All"

                mt = m.get("meeting_type", "General")
                type_colors = {
                    "General": "#3b82f6", "Partner Call": "#10b981",
                    "Training": "#f59e0b", "Workshop": "#8b5cf6",
                    "Review": "#ef4444", "Dissemination": "#ec4899",
                }
                tc = type_colors.get(mt, "#64748b")

                                # ── (page_meetings devamı — upcoming tab içinde) ──

                st.markdown(
                    f"<div class='{card_class}'>"
                    f"<div style='display:flex;justify-content:space-between;align-items:flex-start;'>"
                    f"<div style='flex:1;'>"
                    f"<div class='meeting-date'>📅 {md}{today_badge} "
                    f"<span class='badge' style='background:{tc}20;color:{tc};margin-left:8px;'>{mt}</span></div>"
                    f"<div class='meeting-title'>{m.get('title', 'Untitled')}</div>"
                    f"<div class='meeting-time'>🕐 {str(m.get('start_time',''))[:5]} — "
                    f"{str(m.get('end_time',''))[:5]} ({m.get('timezone','CET')})</div>"
                    f"<div style='margin-top:0.4rem;font-size:0.85rem;color:#64748b;'>👥 {part_str}</div>"
                    f"{'<div style=\"margin-top:0.4rem;font-size:0.9rem;color:#475569;\">' + m.get('description','') + '</div>' if m.get('description') else ''}"
                    f"{'<div style=\"margin-top:0.4rem;\"><strong style=\"font-size:0.85rem;color:#1B3A5C;\">📋 Agenda:</strong><br><span style=\"font-size:0.85rem;color:#475569;white-space:pre-line;\">' + m.get('agenda','') + '</span></div>' if m.get('agenda') else ''}"
                    f"{zoom_info}"
                    f"</div>"
                    f"<div style='text-align:right;min-width:180px;'>{zoom_btn}</div>"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )
            st.markdown(
                f"<p style='text-align:center;color:#94a3b8;margin-top:1rem;'>"
                f"Showing {len(upcoming)} upcoming meeting(s)</p>",
                unsafe_allow_html=True,
            )

    # ─── ALL MEETINGS ───
    with tabs[1]:
        if not meetings:
            st.info("No meetings recorded yet.")
        else:
            st.markdown("### 📋 Complete Meeting History")
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                mt_filter = st.multiselect(
                    "Meeting Type",
                    ["General", "Partner Call", "Training", "Workshop", "Review", "Dissemination"],
                    default=[], key="mt_filter_all",
                )
            with fc2:
                status_filter = st.selectbox("Status", ["All", "Scheduled", "Completed", "Cancelled"], key="st_filter_all")
            with fc3:
                date_range = st.selectbox("Period", ["All", "This Week", "This Month", "Past", "Future"], key="dr_filter_all")

            filtered = meetings.copy()
            if mt_filter:
                filtered = [m for m in filtered if m.get("meeting_type", "General") in mt_filter]
            if status_filter != "All":
                filtered = [m for m in filtered if m.get("status", "Scheduled") == status_filter]
            if date_range == "Past":
                filtered = [m for m in filtered if str(m.get("meeting_date", "")) < str(today)]
            elif date_range == "Future":
                filtered = [m for m in filtered if str(m.get("meeting_date", "")) >= str(today)]
            elif date_range == "This Week":
                week_end = today + timedelta(days=7)
                filtered = [m for m in filtered if str(today) <= str(m.get("meeting_date", "")) <= str(week_end)]
            elif date_range == "This Month":
                month_end = today.replace(day=28) + timedelta(days=4)
                filtered = [m for m in filtered if str(today) <= str(m.get("meeting_date", "")) <= str(month_end)]

            if filtered:
                df_data = []
                for m in filtered:
                    is_past = str(m.get("meeting_date", "")) < str(today)
                    si = "✅" if m.get("status") == "Completed" else ("⏳" if not is_past else "🔘")
                    df_data.append({
                        "Status": si,
                        "Date": str(m.get("meeting_date", ""))[:10],
                        "Time": f"{str(m.get('start_time',''))[:5]}–{str(m.get('end_time',''))[:5]}",
                        "Title": m.get("title", ""),
                        "Type": m.get("meeting_type", "General"),
                        "Zoom": "🔗" if m.get("zoom_link") else "—",
                        "ID": m.get("id", ""),
                    })
                st.dataframe(pd.DataFrame(df_data), use_container_width=True, hide_index=True)

                # Admin management
                if r == "Admin":
                    st.markdown("---")
                    st.markdown("#### ⚙️ Manage Meetings")
                    mc1, mc2 = st.columns(2)
                    not_completed = [m for m in filtered if m.get("status") != "Completed"]
                    with mc1:
                        if not_completed:
                            mid_c = st.selectbox(
                                "Mark as Completed",
                                [f"#{m.get('id')} — {m.get('title','')}" for m in not_completed],
                                key="complete_sel",
                            )
                            if st.button("✅ Mark Completed", key="btn_complete"):
                                mid = int(mid_c.split("#")[1].split(" ")[0])
                                db_update_meeting(mid, {"status": "Completed"})
                                st.success("Marked as completed!")
                                st.rerun()
                    with mc2:
                        if filtered:
                            mid_d = st.selectbox(
                                "Delete Meeting",
                                [f"#{m.get('id')} — {m.get('title','')}" for m in filtered],
                                key="del_sel",
                            )
                            if st.button("🗑️ Delete", key="btn_del"):
                                mid = int(mid_d.split("#")[1].split(" ")[0])
                                db_delete_meeting(mid)
                                st.success("Deleted!")
                                st.rerun()

                    # Notes & Recording
                    st.markdown("#### 📝 Meeting Notes & Recordings")
                    past_m = [m for m in filtered if str(m.get("meeting_date", "")) < str(today) or m.get("status") == "Completed"]
                    if past_m:
                        sel_note = st.selectbox(
                            "Select Meeting",
                            [f"#{m.get('id')} — {m.get('title','')}" for m in past_m],
                            key="note_sel",
                        )
                        if sel_note:
                            mid = int(sel_note.split("#")[1].split(" ")[0])
                            cur_m = next((m for m in past_m if m.get("id") == mid), {})
                            notes = st.text_area("Notes", value=cur_m.get("notes", "") or "", height=120, key="notes_inp")
                            rec_link = st.text_input("Recording Link", value=cur_m.get("recording_link", "") or "", key="rec_inp")
                            if st.button("💾 Save Notes", key="btn_save_notes"):
                                db_update_meeting(mid, {"notes": notes, "recording_link": rec_link})
                                st.success("Saved!")
                                st.rerun()
            else:
                st.info("No meetings match filters.")

    # ─── SCHEDULE (Admin only) ───
    if r == "Admin":
        with tabs[2]:
            st.markdown("### ➕ Schedule New Meeting")
            with st.form("new_meeting"):
                m_title = st.text_input("Meeting Title *", placeholder="e.g. Monthly Partner Coordination Call")
                m_desc = st.text_area("Description", placeholder="Brief description", height=80)

                dc1, dc2, dc3 = st.columns(3)
                with dc1:
                    m_date = st.date_input("Date *", value=today + timedelta(days=7), min_value=today)
                with dc2:
                    m_start = st.time_input("Start Time *", value=dt_time(14, 0))
                with dc3:
                    m_end = st.time_input("End Time *", value=dt_time(15, 0))

                zc1, zc2, zc3 = st.columns(3)
                with zc1:
                    m_zoom = st.text_input("Zoom Link", placeholder="https://zoom.us/j/...")
                with zc2:
                    m_zoom_id = st.text_input("Meeting ID", placeholder="123 456 7890")
                with zc3:
                    m_zoom_pass = st.text_input("Passcode", placeholder="abc123")

                tc1, tc2, tc3 = st.columns(3)
                with tc1:
                    m_type = st.selectbox("Type", ["General", "Partner Call", "Training", "Workshop", "Review", "Dissemination"])
                with tc2:
                    m_tz = st.selectbox("Timezone", ["CET", "EET", "GMT", "UTC"])
                with tc3:
                    m_parts = st.multiselect("Participants", ["All", "Turkey", "Poland", "Spain", "Patients", "External"], default=["All"])

                m_agenda = st.text_area("Agenda", placeholder="1. Welcome\n2. Progress\n3. Next steps", height=120)

                if st.form_submit_button("📅 Schedule Meeting", type="primary", use_container_width=True):
                    if m_title and m_date:
                        db_add_meeting({
                            "title": m_title, "description": m_desc,
                            "meeting_date": str(m_date),
                            "start_time": str(m_start), "end_time": str(m_end),
                            "timezone": m_tz, "zoom_link": m_zoom,
                            "zoom_id": m_zoom_id, "zoom_passcode": m_zoom_pass,
                            "meeting_type": m_type,
                            "participants": json.dumps(m_parts),
                            "agenda": m_agenda, "created_by": get_name(),
                            "status": "Scheduled",
                        })
                        st.success(f"✅ '{m_title}' scheduled for {m_date}!")
                        st.rerun()
                    else:
                        st.error("Title and date required.")

            # Quick templates
            st.markdown("---")
            st.markdown("### 🎯 Quick Templates")
            qt1, qt2, qt3 = st.columns(3)
            with qt1:
                if st.button("📞 Weekly Partner Call", use_container_width=True):
                    next_mon = today + timedelta(days=(7 - today.weekday()))
                    db_add_meeting({
                        "title": "Weekly Partner Coordination Call",
                        "description": "Regular weekly sync.",
                        "meeting_date": str(next_mon),
                        "start_time": "14:00", "end_time": "15:00",
                        "timezone": "CET", "meeting_type": "Partner Call",
                        "participants": json.dumps(["All"]),
                        "agenda": "1. Status updates\n2. Blockers\n3. Next week\n4. AOB",
                        "created_by": get_name(), "status": "Scheduled",
                    })
                    st.success("Scheduled!")
                    st.rerun()
            with qt2:
                if st.button("🎓 Training Session", use_container_width=True):
                    next_wed = today + timedelta(days=(2 - today.weekday()) % 7 + 7)
                    db_add_meeting({
                        "title": "AI Tools Training Session",
                        "description": "Hands-on AI training.",
                        "meeting_date": str(next_wed),
                        "start_time": "10:00", "end_time": "12:00",
                        "timezone": "CET", "meeting_type": "Training",
                        "participants": json.dumps(["All", "Patients"]),
                        "agenda": "1. Overview\n2. Demo\n3. Practice\n4. Q&A",
                        "created_by": get_name(), "status": "Scheduled",
                    })
                    st.success("Scheduled!")
                    st.rerun()
            with qt3:
                if st.button("📊 Monthly Review", use_container_width=True):
                    fon = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
                    db_add_meeting({
                        "title": "Monthly Progress Review",
                        "description": "Monthly review of progress, deliverables, budget.",
                        "meeting_date": str(fon),
                        "start_time": "14:00", "end_time": "16:00",
                        "timezone": "CET", "meeting_type": "Review",
                        "participants": json.dumps(["All"]),
                        "agenda": "1. WP progress\n2. Budget\n3. Deliverables\n4. Risks\n5. Plan",
                        "created_by": get_name(), "status": "Scheduled",
                    })
                    st.success("Scheduled!")
                    st.rerun()

    # ─── CALENDAR VIEW ───
    cal_tab_idx = 3 if r == "Admin" else 2
    with tabs[cal_tab_idx]:
        st.markdown("### 🗓️ Calendar Overview")
        if meetings:
            cal_data = []
            for m in meetings:
                md = str(m.get("meeting_date", ""))
                try:
                    d_obj = datetime.strptime(md[:10], "%Y-%m-%d")
                except Exception:
                    continue
                cal_data.append({
                    "Date": d_obj,
                    "Title": m.get("title", ""),
                    "Type": m.get("meeting_type", "General"),
                    "Time": str(m.get("start_time", ""))[:5],
                    "Status": m.get("status", "Scheduled"),
                })
            if cal_data:
                df_cal = pd.DataFrame(cal_data)
                df_cal["Month"] = df_cal["Date"].dt.strftime("%Y-%m")

                fig = px.scatter(
                    df_cal, x="Date", y="Type", color="Type", size_max=15,
                    hover_data=["Title", "Time", "Status"],
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                fig.update_traces(marker=dict(size=14, symbol="diamond"))
                fig.update_layout(height=400, title="Meeting Timeline", yaxis_title="", xaxis_title="")
                fig.add_vline(x=datetime.now(), line_dash="dash", line_color="red", annotation_text="Today")
                st.plotly_chart(fig, use_container_width=True)

                monthly = df_cal.groupby("Month").size().reset_index(name="Count")
                fig2 = px.bar(monthly, x="Month", y="Count", color_discrete_sequence=["#2ABFBF"])
                fig2.update_layout(height=300, title="Meetings per Month")
                st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No meetings to display.")

        # Stats block
        st.markdown("### 📊 Meeting Statistics")
        total = len(meetings)
        completed = len([m for m in meetings if m.get("status") == "Completed"])
        upcoming_c = len([m for m in meetings if str(m.get("meeting_date", "")) >= str(today)])
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            st.markdown(f"<div class='metric-card'><div class='value'>{total}</div><div class='label'>Total</div></div>", unsafe_allow_html=True)
        with sc2:
            st.markdown(f"<div class='metric-card'><div class='value'>{completed}</div><div class='label'>Completed</div></div>", unsafe_allow_html=True)
        with sc3:
            st.markdown(f"<div class='metric-card'><div class='value'>{upcoming_c}</div><div class='label'>Upcoming</div></div>", unsafe_allow_html=True)

        if meetings:
            types = {}
            for m in meetings:
                t = m.get("meeting_type", "General")
                types[t] = types.get(t, 0) + 1
            fig3 = px.pie(values=list(types.values()), names=list(types.keys()), color_discrete_sequence=px.colors.qualitative.Pastel)
            fig3.update_layout(height=350, title="By Type")
            st.plotly_chart(fig3, use_container_width=True)


# ═══════════════════════════════════════
# PAGE: WORK PACKAGES
# ═══════════════════════════════════════
def page_work_packages(wp_df):
    r = get_role()
    c = get_country()
    st.markdown(
        "<div class='pro-header'><h1>📦 Work Packages</h1>"
        "<p>Project structure and task allocation</p></div>",
        unsafe_allow_html=True,
    )
    if len(wp_df) == 0:
        st.info("No WP data. Create data/work_packages.csv")
        return

    # Filter for partner role
    if r == "Partner" and c != "All":
        wps = get_user_wps(c)
        wp_df = wp_df[wp_df["wp_id"].isin(wps)]
        st.info(f"Showing WPs for {FLAGS.get(c,'')} {c}: {', '.join(wps)}")

    # Summary metrics
    total_budget = wp_df["budget_eur"].sum() if "budget_eur" in wp_df.columns else 0
    in_progress = len(wp_df[wp_df["status"] == "In Progress"]) if "status" in wp_df.columns else 0
    not_started = len(wp_df[wp_df["status"] == "Not Started"]) if "status" in wp_df.columns else 0

    mc1, mc2, mc3, mc4 = st.columns(4)
    with mc1:
        st.markdown(f"<div class='metric-card'><div class='value'>{len(wp_df)}</div><div class='label'>Work Packages</div></div>", unsafe_allow_html=True)
    with mc2:
        st.markdown(f"<div class='metric-card'><div class='value'>€{total_budget:,.0f}</div><div class='label'>Total Budget</div></div>", unsafe_allow_html=True)
    with mc3:
        st.markdown(f"<div class='metric-card'><div class='value'>{in_progress}</div><div class='label'>In Progress</div></div>", unsafe_allow_html=True)
    with mc4:
        st.markdown(f"<div class='metric-card'><div class='value'>{not_started}</div><div class='label'>Not Started</div></div>", unsafe_allow_html=True)

    st.markdown("")

    for _, row in wp_df.iterrows():
        wpid = row.get("wp_id", "")
        wp_name = row.get("wp_name", row.get("title", "Untitled"))
        lead = row.get("lead_partner", row.get("lead_country", "TBD"))
        lead_country = row.get("lead_country", "TBD")
        status = row.get("status", "Planned")
        budget = row.get("budget_eur", 0)
        start_m = row.get("start_month", "")
        end_m = row.get("end_month", "")
        desc = row.get("description", "N/A")
        deliverables = row.get("deliverables", "")
        supporting = row.get("supporting_partners", "")

        # Status color
        status_colors = {
            "Completed": "badge-success",
            "In Progress": "badge-info",
            "Not Started": "badge-warning",
            "Planned": "badge-purple" if "badge-purple" in "" else "badge-warning",
        }
        status_badge = status_colors.get(status, "badge-info")

        # Calculate actual dates from months
        if start_m and end_m:
            try:
                s_date = PROJECT_START + timedelta(days=(int(start_m) - 1) * 30)
                e_date = PROJECT_START + timedelta(days=int(end_m) * 30)
                date_str = f"{s_date.strftime('%b %Y')} → {e_date.strftime('%b %Y')} (M{start_m}–M{end_m})"
            except Exception:
                date_str = f"M{start_m} – M{end_m}"
        else:
            date_str = "TBD"

        with st.expander(f"📦 {wpid}: {wp_name}", expanded=False):
            # Top metrics
            tc1, tc2, tc3, tc4 = st.columns(4)
            with tc1:
                st.metric("Budget", f"€{budget:,.0f}")
            with tc2:
                st.metric("Status", status)
            with tc3:
                st.metric("Lead", lead_country)
            with tc4:
                st.metric("Timeline", f"M{start_m}–M{end_m}")

            st.markdown(f"**📅 Period:** {date_str}")
            st.markdown(f"**🏢 Lead Organisation:** {lead}")

            # Supporting partners
            if supporting:
                sup_list = supporting.split(";")
                st.markdown(f"**🤝 Supporting:** {', '.join(sup_list)}")

            # Country roles
            if r != "Patient":
                st.markdown("**🌍 Country Roles:**")
                rc = st.columns(3)
                for col2, cn in zip(rc, ["Turkey", "Poland", "Spain"]):
                    with col2:
                        st.markdown(f"{FLAGS[cn]} **{cn}**: {get_wp_role(cn, wpid)}")

            st.markdown(f"**📝 Description:** {desc}")

            # Deliverables
            if deliverables:
                st.markdown("**📦 Deliverables:**")
                for d in deliverables.split(";"):
                    d = d.strip()
                    if d:
                        st.markdown(f"  - 📄 {d}")

    # Budget overview chart
    if "budget_eur" in wp_df.columns:
        st.markdown("---")
        st.subheader("💰 Budget Overview")
        label_col = "wp_name" if "wp_name" in wp_df.columns else "wp_id"
        fig = px.bar(
            wp_df, x="wp_id", y="budget_eur", color="status",
            text="budget_eur", hover_data=[label_col],
            color_discrete_map={
                "Completed": "#10B981",
                "In Progress": "#3B82F6",
                "Not Started": "#94A3B8",
            },
        )
        fig.update_traces(texttemplate="€%{text:,.0f}", textposition="outside")
        fig.update_layout(height=400, yaxis_title="Budget (EUR)", xaxis_title="")
        st.plotly_chart(fig, use_container_width=True)


# ═══════════════════════════════════════
# PAGE: GANTT CHART
# ═══════════════════════════════════════
def page_gantt(wp_df):
    r = get_role()
    c = get_country()
    st.markdown(
        "<div class='pro-header'><h1>📅 Interactive Gantt Chart</h1>"
        "<p>Dynamic project timeline — zoom, filter, track progress</p></div>",
        unsafe_allow_html=True,
    )
    if len(wp_df) == 0:
        st.info("No WP data. Create data/work_packages.csv")
        return

    if "start_month" not in wp_df.columns or "end_month" not in wp_df.columns:
        st.warning("CSV must have start_month and end_month columns.")
        return

    df = wp_df.copy()

    # Partner filtering
    if r == "Partner" and c != "All":
        wps = get_user_wps(c)
        df = df[df["wp_id"].isin(wps)]

    # Convert months to dates
    try:
        df["start_date"] = df["start_month"].apply(
            lambda m: PROJECT_START + timedelta(days=(int(m) - 1) * 30)
        )
        df["end_date"] = df["end_month"].apply(
            lambda m: PROJECT_START + timedelta(days=int(m) * 30)
        )
    except Exception as e:
        st.error(f"Date conversion error: {e}")
        return

    label_col = "wp_name" if "wp_name" in df.columns else "wp_id"
    df["label"] = df["wp_id"] + ": " + df[label_col]
    df["duration_months"] = df["end_month"].astype(int) - df["start_month"].astype(int) + 1

    # Calculate progress for each WP
    now = datetime.now()
    progress_list = []
    for _, row in df.iterrows():
        sd = row["start_date"]
        ed = row["end_date"]
        if now < sd:
            progress_list.append(0)
        elif now > ed:
            progress_list.append(100)
        else:
            total = (ed - sd).days
            elapsed = (now - sd).days
            progress_list.append(int((elapsed / max(total, 1)) * 100))
    df["progress"] = progress_list

    if df.empty:
        st.warning("No data to display.")
        return

    # ═══════════════════════════════════════
    # INTERACTIVE CONTROLS
    # ═══════════════════════════════════════
    st.markdown("### 🎛️ Controls")
    ctrl1, ctrl2, ctrl3, ctrl4 = st.columns(4)

    with ctrl1:
        view_mode = st.selectbox(
            "📊 View",
            ["Timeline", "Progress", "Budget", "Country Lead", "All-in-One"],
            key="gantt_view",
        )
    with ctrl2:
        status_filter = st.multiselect(
            "🔖 Status",
            df["status"].unique().tolist(),
            default=df["status"].unique().tolist(),
            key="gantt_status",
        )
    with ctrl3:
        if "lead_country" in df.columns:
            country_filter = st.multiselect(
                "🌍 Country",
                df["lead_country"].unique().tolist(),
                default=df["lead_country"].unique().tolist(),
                key="gantt_country",
            )
        else:
            country_filter = []
    with ctrl4:
        show_milestones = st.checkbox("📍 Show Milestones", value=True, key="gantt_miles")

    # Apply filters
    filtered = df.copy()
    if status_filter:
        filtered = filtered[filtered["status"].isin(status_filter)]
    if country_filter and "lead_country" in filtered.columns:
        filtered = filtered[filtered["lead_country"].isin(country_filter)]

    if filtered.empty:
        st.warning("No WPs match the selected filters.")
        return

    # ═══════════════════════════════════════
    # VIEW 1: TIMELINE (Default Gantt)
    # ═══════════════════════════════════════
    if view_mode in ("Timeline", "All-in-One"):
        st.markdown("### 📅 Project Timeline")

        fig = go.Figure()

        status_colors = {
            "Completed": "#10B981",
            "In Progress": "#3B82F6",
            "Not Started": "#94A3B8",
            "Planned": "#F59E0B",
        }

        for idx, row in filtered.iterrows():
            color = status_colors.get(row["status"], "#64748b")
            prog = row["progress"]

            # Full bar (background)
            fig.add_trace(go.Bar(
                x=[(row["end_date"] - row["start_date"]).days],
                y=[row["label"]],
                base=[row["start_date"]],
                orientation="h",
                marker=dict(
                    color=color,
                    opacity=0.3,
                    line=dict(width=1, color=color),
                ),
                hovertemplate=(
                    f"<b>{row['label']}</b><br>"
                    f"Lead: {row.get('lead_country', 'N/A')}<br>"
                    f"Period: M{row['start_month']}–M{row['end_month']} "
                    f"({row['duration_months']} months)<br>"
                    f"Budget: €{row.get('budget_eur', 0):,.0f}<br>"
                    f"Status: {row['status']}<br>"
                    f"Progress: {prog}%<br>"
                    f"<extra></extra>"
                ),
                showlegend=False,
                name=row["wp_id"],
            ))

            # Progress bar (foreground)
            if prog > 0:
                total_days = (row["end_date"] - row["start_date"]).days
                prog_days = int(total_days * prog / 100)
                fig.add_trace(go.Bar(
                    x=[prog_days],
                    y=[row["label"]],
                    base=[row["start_date"]],
                    orientation="h",
                    marker=dict(color=color, opacity=0.85),
                    showlegend=False,
                    hoverinfo="skip",
                ))

            # Progress text
            mid_date = row["start_date"] + timedelta(
                days=(row["end_date"] - row["start_date"]).days // 2
            )
            fig.add_annotation(
                x=mid_date,
                y=row["label"],
                text=f"{prog}%",
                showarrow=False,
                font=dict(color="white" if prog > 30 else "#333", size=11, family="Inter"),
            )

        # Today marker
        fig.add_shape(
            type="line",
            x0=now, x1=now,
            y0=-0.5, y1=len(filtered) - 0.5,
            line=dict(color="#EF4444", width=2.5, dash="dash"),
        )
        fig.add_annotation(
            x=now, y=len(filtered) - 0.3,
            text="📍 TODAY",
            showarrow=True,
            arrowhead=2,
            arrowcolor="#EF4444",
            font=dict(color="#EF4444", size=11, family="Inter, sans-serif"),
            bgcolor="white",
            bordercolor="#EF4444",
            borderwidth=1,
            borderpad=4,
            ax=0, ay=-30,
        )

        if show_milestones:
            # Project Start
            fig.add_shape(
                type="line",
                x0=PROJECT_START, x1=PROJECT_START,
                y0=-0.5, y1=len(filtered) - 0.5,
                line=dict(color="#2ABFBF", width=1.5, dash="dot"),
            )
            fig.add_annotation(
                x=PROJECT_START, y=-0.7,
                text="🚀 Start", showarrow=False,
                font=dict(color="#2ABFBF", size=10),
            )

            # Project End
            project_end = PROJECT_START + timedelta(days=18 * 30)
            fig.add_shape(
                type="line",
                x0=project_end, x1=project_end,
                y0=-0.5, y1=len(filtered) - 0.5,
                line=dict(color="#dc3545", width=1.5, dash="dot"),
            )
            fig.add_annotation(
                x=project_end, y=-0.7,
                text="🏁 End", showarrow=False,
                font=dict(color="#dc3545", size=10),
            )

            # Mid-term review (M9)
            mid_review = PROJECT_START + timedelta(days=9 * 30)
            fig.add_shape(
                type="line",
                x0=mid_review, x1=mid_review,
                y0=-0.5, y1=len(filtered) - 0.5,
                line=dict(color="#F59E0B", width=1, dash="dashdot"),
            )
            fig.add_annotation(
                x=mid_review, y=-0.7,
                text="📋 Mid-Review", showarrow=False,
                font=dict(color="#F59E0B", size=9),
            )

        fig.update_layout(
            height=max(400, len(filtered) * 80 + 100),
            barmode="overlay",
            xaxis=dict(
                title="",
                type="date",
                tickformat="%b %Y",
                dtick="M1",
                tickangle=45,
                gridcolor="#f0f0f0",
                range=[
                    PROJECT_START - timedelta(days=15),
                    PROJECT_START + timedelta(days=18 * 30 + 30),
                ],
                rangeslider=dict(visible=True, thickness=0.05),
            ),
            yaxis=dict(title="", autorange="reversed"),
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(family="Inter, sans-serif"),
            margin=dict(l=10, r=10, t=40, b=80),
            dragmode="zoom",
        )

        config = {
            "displayModeBar": True,
            "modeBarButtonsToAdd": ["drawline", "eraseshape"],
            "scrollZoom": True,
            "displaylogo": False,
        }
        st.plotly_chart(fig, use_container_width=True, config=config)

    # ═══════════════════════════════════════
    # VIEW 2: PROGRESS TRACKER
    # ═══════════════════════════════════════
    if view_mode in ("Progress", "All-in-One"):
        st.markdown("### 📈 Progress Tracker")

        prog_cols = st.columns(len(filtered))
        for col_p, (_, row) in zip(prog_cols, filtered.iterrows()):
            with col_p:
                prog = row["progress"]
                if prog >= 100:
                    color = "#10B981"
                    emoji = "✅"
                elif prog >= 50:
                    color = "#3B82F6"
                    emoji = "🔄"
                elif prog > 0:
                    color = "#F59E0B"
                    emoji = "🔶"
                else:
                    color = "#94A3B8"
                    emoji = "⏳"

                st.markdown(
                    f"<div style='text-align:center;padding:1rem;background:white;"
                    f"border-radius:12px;border:1px solid #e8ecf1;"
                    f"box-shadow:0 2px 8px rgba(0,0,0,0.04);'>"
                    f"<div style='font-size:1.5rem;'>{emoji}</div>"
                    f"<div style='font-weight:700;color:#1B3A5C;font-size:0.9rem;'>{row['wp_id']}</div>"
                    f"<div style='font-size:0.75rem;color:#64748b;margin:0.3rem 0;'>"
                    f"{row.get('wp_name', '')[:20]}</div>"
                    f"<div style='background:#e2e8f0;border-radius:10px;height:12px;"
                    f"margin:0.5rem 0;overflow:hidden;'>"
                    f"<div style='background:{color};height:100%;width:{prog}%;"
                    f"border-radius:10px;transition:width 0.5s;'></div></div>"
                    f"<div style='font-size:1.3rem;font-weight:800;color:{color};'>{prog}%</div>"
                    f"<div style='font-size:0.7rem;color:#94a3b8;'>{row['status']}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        # Overall progress
        avg_prog = int(filtered["progress"].mean())
        st.markdown(
            f"<div style='background:linear-gradient(135deg,#1B3A5C,#2d5a8e);border-radius:16px;"
            f"padding:1.5rem;color:white;text-align:center;margin:1rem 0;'>"
            f"<p style='margin:0;font-size:0.8rem;opacity:0.7;letter-spacing:1px;'>OVERALL PROJECT PROGRESS</p>"
            f"<div style='background:rgba(255,255,255,0.2);border-radius:10px;height:20px;"
            f"margin:1rem 2rem;overflow:hidden;'>"
            f"<div style='background:#2ABFBF;height:100%;width:{avg_prog}%;border-radius:10px;"
            f"transition:width 0.5s;'></div></div>"
            f"<p style='margin:0;font-size:2rem;font-weight:800;'>{avg_prog}%</p></div>",
            unsafe_allow_html=True,
        )

    # ═══════════════════════════════════════
    # VIEW 3: BUDGET TIMELINE
    # ═══════════════════════════════════════
    if view_mode in ("Budget", "All-in-One"):
        st.markdown("### 💰 Budget Distribution Timeline")

        fig_b = px.timeline(
            filtered,
            x_start="start_date",
            x_end="end_date",
            y="label",
            color="budget_eur",
            color_continuous_scale="Viridis",
            hover_data=["lead_country", "status", "budget_eur", "duration_months"],
        )
        fig_b.update_yaxes(autorange="reversed")
        fig_b.update_layout(
            height=max(350, len(filtered) * 65),
            coloraxis_colorbar=dict(title="Budget €"),
            xaxis=dict(tickformat="%b %Y", dtick="M1", tickangle=45),
            plot_bgcolor="white",
        )
        st.plotly_chart(fig_b, use_container_width=True)

        # Budget bar chart
        fig_bar = px.bar(
            filtered, x="wp_id", y="budget_eur", color="lead_country",
            text="budget_eur", title="Budget per WP",
            color_discrete_map={
                "Turkey": "#E30A17", "Poland": "#DC143C", "Spain": "#F4B400",
            },
        )
        fig_bar.update_traces(texttemplate="€%{text:,.0f}", textposition="outside")
        fig_bar.update_layout(height=350, yaxis_title="EUR", xaxis_title="")
        st.plotly_chart(fig_bar, use_container_width=True)

    # ═══════════════════════════════════════
    # VIEW 4: COUNTRY LEAD VIEW
    # ═══════════════════════════════════════
    if view_mode in ("Country Lead", "All-in-One"):
        st.markdown("### 🌍 Country Lead Overview")

        if "lead_country" in filtered.columns:
            fig_c = px.timeline(
                filtered,
                x_start="start_date",
                x_end="end_date",
                y="label",
                color="lead_country",
                color_discrete_map={
                    "Turkey": "#E30A17",
                    "Poland": "#DC143C",
                    "Spain": "#F4B400",
                },
                hover_data=["status", "budget_eur"],
            )
            fig_c.update_yaxes(autorange="reversed")
            fig_c.update_layout(
                height=max(350, len(filtered) * 65),
                xaxis=dict(tickformat="%b %Y", dtick="M1", tickangle=45),
                plot_bgcolor="white",
            )

            # Today
            fig_c.add_shape(
                type="line",
                x0=now, x1=now,
                y0=-0.5, y1=len(filtered) - 0.5,
                line=dict(color="#EF4444", width=2, dash="dash"),
            )

            st.plotly_chart(fig_c, use_container_width=True)

            # Country summary cards
            cc = st.columns(3)
            for col_cc, cn in zip(cc, ["Turkey", "Poland", "Spain"]):
                with col_cc:
                    cn_wps = filtered[filtered["lead_country"] == cn]
                    cn_budget = cn_wps["budget_eur"].sum() if len(cn_wps) > 0 else 0
                    cn_count = len(cn_wps)
                    st.markdown(
                        f"<div style='background:white;border-radius:12px;padding:1.2rem;"
                        f"text-align:center;border:1px solid #e8ecf1;'>"
                        f"<div style='font-size:2rem;'>{FLAGS.get(cn, '🌍')}</div>"
                        f"<div style='font-weight:700;color:#1B3A5C;'>{cn}</div>"
                        f"<div style='color:#2ABFBF;font-size:1.3rem;font-weight:800;'>"
                        f"€{cn_budget:,.0f}</div>"
                        f"<div style='color:#94a3b8;font-size:0.85rem;'>{cn_count} WP(s) lead</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

    # ═══════════════════════════════════════
    # MONTHLY HEATMAP (always shown)
    # ═══════════════════════════════════════
    st.markdown("---")
    st.markdown("### 📊 Monthly Activity Heatmap")

    months = list(range(1, 19))
    month_labels = []
    for m in months:
        d = PROJECT_START + timedelta(days=(m - 1) * 30)
        month_labels.append(f"M{m}\n{d.strftime('%b %y')}")

    heatmap_data = []
    for _, row in filtered.iterrows():
        sm = int(row.get("start_month", 1))
        em = int(row.get("end_month", 18))
        for m in months:
            active = 1 if sm <= m <= em else 0
            heatmap_data.append({
                "WP": row.get("wp_id", ""),
                "Month": f"M{m}",
                "Active": active,
            })

    hm_df = pd.DataFrame(heatmap_data)
    pivot = hm_df.pivot(index="WP", columns="Month", values="Active")

    # Sort columns properly
    sorted_cols = [f"M{i}" for i in range(1, 19)]
    pivot = pivot[[c2 for c2 in sorted_cols if c2 in pivot.columns]]

    fig_hm = go.Figure(data=go.Heatmap(
        z=pivot.values,
        x=pivot.columns,
        y=pivot.index,
        colorscale=[[0, "#f1f5f9"], [1, "#2ABFBF"]],
        showscale=False,
        hovertemplate="<b>%{y}</b><br>%{x}<br>Active: %{z}<extra></extra>",
    ))

    # Current month marker
    elapsed_months = max(1, min(18, ((now - PROJECT_START).days // 30) + 1))
    fig_hm.add_shape(
        type="rect",
        x0=elapsed_months - 1.5, x1=elapsed_months - 0.5,
        y0=-0.5, y1=len(pivot) - 0.5,
        line=dict(color="#EF4444", width=2),
        fillcolor="rgba(239,68,68,0.1)",
    )

    fig_hm.update_layout(
        height=max(250, len(filtered) * 45 + 80),
        xaxis=dict(title="", tickangle=0, side="top"),
        yaxis=dict(title="", autorange="reversed"),
        plot_bgcolor="white",
        margin=dict(l=10, r=10, t=40, b=20),
    )
    st.plotly_chart(fig_hm, use_container_width=True)

    # ═══════════════════════════════════════
    # TIMELINE DETAILS TABLE
    # ═══════════════════════════════════════
    st.markdown("### 📋 Timeline Details")
    timeline_data = []
    for _, row in filtered.iterrows():
        prog = row["progress"]
        if prog >= 100:
            prog_icon = "✅"
        elif prog >= 50:
            prog_icon = "🔄"
        elif prog > 0:
            prog_icon = "🔶"
        else:
            prog_icon = "⏳"

        timeline_data.append({
            "": prog_icon,
            "WP": row.get("wp_id", ""),
            "Name": row.get("wp_name", ""),
            "Lead": f"{FLAGS.get(row.get('lead_country', ''), '🌍')} {row.get('lead_country', '')}",
            "Start": f"M{row.get('start_month', '')} ({row['start_date'].strftime('%b %Y')})",
            "End": f"M{row.get('end_month', '')} ({row['end_date'].strftime('%b %Y')})",
            "Duration": f"{row['duration_months']} mo",
            "Progress": f"{prog}%",
            "Status": row.get("status", ""),
            "Budget": f"€{row.get('budget_eur', 0):,.0f}",
        })
    st.dataframe(
        pd.DataFrame(timeline_data),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Progress": st.column_config.ProgressColumn(
                "Progress", min_value=0, max_value=100,
                format="%d%%",
            ),
        },
    )

# ═══════════════════════════════════════
# PAGE: PARTNERS
# ═══════════════════════════════════════
def page_partners(pf):
    st.markdown(
        "<div class='pro-header'><h1>🤝 Partner Organisations</h1>"
        "<p>Consortium members and responsibilities</p></div>",
        unsafe_allow_html=True,
    )
    for cn, org in PARTNER_MAP.items():
        with st.expander(f"{FLAGS[cn]} {org} — {cn}", expanded=True):
            st.markdown(f"**Country:** {cn}")
            st.markdown(f"**Organisation:** {org}")
            wpm = WP_COUNTRY_MAP.get(cn, {})
            st.markdown(f"**Lead WPs:** {', '.join(wpm.get('lead', []))}")
            st.markdown(f"**Support WPs:** {', '.join(wpm.get('support', []))}")


# ═══════════════════════════════════════
# PAGE: PARTNER FEEDBACK
# ═══════════════════════════════════════
def page_partner_feedback():
    r = get_role()
    c = get_country()
    st.markdown(
        "<div class='pro-header'><h1>💬 Partner Feedback</h1>"
        "<p>Submit and manage proposal feedback</p></div>",
        unsafe_allow_html=True,
    )
    perm = get_permission("Partner Feedback", r)
    if perm in ("write", "full"):
        with st.expander("➕ Submit New Feedback", expanded=False):
            with st.form("pfb_form"):
                sec = st.selectbox("Section", PROPOSAL_SECTIONS)
                fb = st.text_area("Feedback", height=120)
                pri = st.select_slider("Priority", ["Low", "Medium", "High"], value="Medium")
                if st.form_submit_button("📤 Submit", type="primary", use_container_width=True):
                    if fb.strip():
                        cn = c if r == "Partner" else "Admin"
                        db_add_partner_feedback(cn, get_org(), sec, fb.strip(), pri, get_name())
                        st.success("✅ Submitted!")
                        st.rerun()
                    else:
                        st.warning("Please write feedback.")

    fbl = db_get_partner_feedback()
    if r == "Partner" and c != "All":
        fbl = [f for f in fbl if f.get("partner_country", f.get("country", "")) in (c, "Admin")]

    if not fbl:
        st.info("No feedback yet.")
        return

    # Export
    if EXPORT_OK and r == "Admin":
        ec1, ec2, _ = st.columns([1, 1, 4])
        with ec1:
            xl = generate_feedback_excel()
            if xl:
                st.download_button("📥 Excel", xl, "feedback.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with ec2:
            pdf = generate_feedback_pdf()
            if pdf:
                st.download_button("📥 PDF", pdf, "feedback.pdf", "application/pdf", use_container_width=True)

    # Filters
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        sf = st.multiselect("Section", sorted(set(f.get("section", "") for f in fbl)), key="pfb_sf")
    with fc2:
        pf2 = st.multiselect("Priority", ["High", "Medium", "Low"], key="pfb_pf")
    with fc3:
        stf = st.multiselect("Status", sorted(set(f.get("status", "Open") for f in fbl)), key="pfb_stf")

    if sf:
        fbl = [f for f in fbl if f.get("section", "") in sf]
    if pf2:
        fbl = [f for f in fbl if f.get("priority", "") in pf2]
    if stf:
        fbl = [f for f in fbl if f.get("status", "Open") in stf]

    st.markdown(f"**{len(fbl)} feedback item(s)**")

    for fb in fbl:
        pri = fb.get("priority", "Medium")
        bc = {"High": "badge-danger", "Medium": "badge-warning", "Low": "badge-success"}.get(pri, "badge-info")
        sc2 = fb.get("status", "Open")
        sbc = {"Open": "badge-info", "Accepted": "badge-success", "Rejected": "badge-danger", "In Review": "badge-warning"}.get(sc2, "badge-info")
        fc = fb.get("partner_country", fb.get("country", ""))

        st.markdown(
            f"<div class='fb-card'>"
            f"<div style='display:flex;justify-content:space-between;align-items:center;'>"
            f"<div><strong>{FLAGS.get(fc,'')} #{fb.get('id','')} — {fb.get('section','')}</strong></div>"
            f"<div><span class='badge {bc}'>{pri}</span> <span class='badge {sbc}'>{sc2}</span></div></div>"
            f"<p style='margin:0.5rem 0;color:#475569;'>{fb.get('feedback', fb.get('content',''))}</p>"
            f"<span style='font-size:0.8rem;color:#94a3b8;'>By {fb.get('submitted_by','')} | {str(fb.get('created_at',''))[:10]}</span>"
            f"{'<div style=\"margin-top:0.5rem;padding:0.5rem;background:#f0fdf4;border-radius:8px;font-size:0.85rem;\"><strong>Response:</strong> ' + fb.get('response','') + '</div>' if fb.get('response') else ''}"
            f"</div>",
            unsafe_allow_html=True,
        )

        if r == "Admin" and sc2 == "Open":
            uc1, uc2, uc3 = st.columns([2, 1, 1])
            with uc1:
                resp = st.text_input("Response", key=f"resp_{fb.get('id')}", placeholder="Admin response...")
            with uc2:
                if st.button("✅ Accept", key=f"acc_{fb.get('id')}"):
                    db_update_feedback_status(fb.get("id"), "Accepted", resp)
                    st.rerun()
            with uc3:
                if st.button("❌ Reject", key=f"rej_{fb.get('id')}"):
                    db_update_feedback_status(fb.get("id"), "Rejected", resp)
                    st.rerun()


# ═══════════════════════════════════════
# PAGE: PATIENT FEEDBACK
# ═══════════════════════════════════════
def page_patient_feedback():
    r = get_role()
    st.markdown(
        "<div class='pro-header'><h1>💚 Patient Feedback</h1>"
        "<p>Patient experience and needs assessment</p></div>",
        unsafe_allow_html=True,
    )
    if r == "Patient":
        with st.form("patient_fb"):
            st.markdown("#### Share Your Experience")
            c1, c2 = st.columns(2)
            with c1:
                country = st.selectbox("Country", ["Turkey", "Poland", "Spain", "Other"])
                age = st.selectbox("Age Group", ["18-30", "31-45", "46-60", "60+"])
            with c2:
                cancer = st.text_input("Cancer Type", placeholder="e.g. Breast, Lung...")
                support = st.selectbox("Support Need", ["Emotional", "Medical Info", "Practical", "Social", "All"])
            digital = st.select_slider("Digital Literacy", ["Low", "Medium", "High"], value="Medium")
            matching = st.selectbox("Matching Preference", ["Same cancer type", "Same age group", "Same country", "No preference"])
            privacy = st.selectbox("Privacy Expectation", ["Anonymous", "First name only", "Full profile OK"])
            comments = st.text_area("Additional Comments", height=100)
            if st.form_submit_button("💚 Submit", type="primary", use_container_width=True):
                db_add_patient_feedback({
                    "country": country, "age_group": age, "cancer_type": cancer,
                    "support_need": support, "digital_literacy": digital,
                    "matching_preference": matching, "privacy_expectation": privacy,
                    "comments": comments, "submitted_by": get_name(),
                })
                st.success("Thank you! 💚")

    if r == "Admin" or st.session_state.get("can_read_patient_fb"):
        pfb = db_get_patient_feedback()
        if pfb:
            st.subheader(f"📊 Patient Feedback ({len(pfb)} responses)")
            df = pd.DataFrame(pfb)
            if "support_need" in df.columns:
                fig = px.pie(df, names="support_need", title="Support Needs", color_discrete_sequence=px.colors.qualitative.Pastel)
                st.plotly_chart(fig, use_container_width=True)
            show_cols = [c for c in ["country", "age_group", "cancer_type", "support_need", "digital_literacy", "matching_preference", "privacy_expectation", "comments", "created_at"] if c in df.columns]
            st.dataframe(df[show_cols], use_container_width=True, hide_index=True)
        else:
            st.info("No patient feedback yet.")
    elif r != "Patient":
        show_access_denied()


# ═══════════════════════════════════════
# PAGE: APPROVAL STATUS
# ═══════════════════════════════════════
def page_approval():
    r = get_role()
    c = get_country()
    st.markdown(
        "<div class='pro-header'><h1>🗳️ Approval Status</h1>"
        "<p>Partner approval tracking</p></div>",
        unsafe_allow_html=True,
    )
    ap = db_get_approvals()
    cols = st.columns(3)
    for col, cn in zip(cols, ["Turkey", "Poland", "Spain"]):
        with col:
            approved = ap.get(cn, False)
            if approved:
                bg = "linear-gradient(135deg,#dcfce7,#f0fdf4)"
                border_color = "#bbf7d0"
                ic = "✅"
                status_text = "Approved"
            else:
                bg = "linear-gradient(135deg,#fef9c3,#fffbeb)"
                border_color = "#fde68a"
                ic = "⏳"
                status_text = "Pending"

            st.markdown(
                f"<div style='background:{bg};border-radius:16px;padding:2rem;text-align:center;"
                f"border:1px solid {border_color};'>"
                f"<div style='font-size:3rem;'>{FLAGS[cn]}</div>"
                f"<h3 style='margin:0.5rem 0;'>{PARTNER_MAP[cn]}</h3>"
                f"<div style='font-size:2rem;'>{ic}</div>"
                f"<p style='color:#64748b;'>{status_text}</p></div>",
                unsafe_allow_html=True,
            )

    if r == "Partner" and c in PARTNER_MAP:
        st.markdown("---")
        if ap.get(c):
            if st.button(f"🔄 Revoke {c} Approval", type="secondary"):
                db_set_approval(c, False, get_name(), r)
                st.rerun()
        else:
            if st.button(f"✅ Approve on behalf of {c}", type="primary"):
                db_set_approval(c, True, get_name(), r)
                st.rerun()

    if r == "Admin":
        st.markdown("---")
        st.subheader("🛡️ Admin Controls")
        for col2, cn in zip(st.columns(3), ["Turkey", "Poland", "Spain"]):
            with col2:
                if ap.get(cn):
                    if st.button(f"Revoke {cn}", key=f"rev_{cn}"):
                        db_set_approval(cn, False, get_name(), "Admin")
                        st.rerun()
                else:
                    if st.button(f"Approve {cn}", key=f"apv_{cn}"):
                        db_set_approval(cn, True, get_name(), "Admin")
                        st.rerun()
        if st.button("🔄 Reset All", type="secondary"):
            db_reset_all_approvals(get_name())
            st.rerun()

    log = db_get_approval_log()
    if log:
        st.subheader("📋 History")
        df_log = pd.DataFrame(log)
        show_cols = [c2 for c2 in ["action", "country", "performed_by", "role", "created_at"] if c2 in df_log.columns]
        st.dataframe(df_log[show_cols].head(20), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════
# PAGE: ANNOUNCEMENTS
# ═══════════════════════════════════════
def page_announcements():
    r = get_role()
    st.markdown(
        "<div class='pro-header'><h1>📢 Announcements</h1>"
        "<p>Project updates and important notices</p></div>",
        unsafe_allow_html=True,
    )
    perm = get_permission("Announcements", r)
    if perm in ("write", "full"):
        with st.expander("➕ Post Announcement", expanded=False):
            with st.form("ann_form"):
                title = st.text_input("Title")
                content = st.text_area("Content", height=100)
                pri = st.select_slider("Priority", ["Low", "Medium", "High"], value="Medium")
                if st.form_submit_button("📢 Post", type="primary"):
                    if title and content:
                        db_add_announcement(title, content, get_name(), pri)
                        st.success("Posted!")
                        st.rerun()
    for a in db_get_announcements():
        ann_card(a)


# ═══════════════════════════════════════
# PAGE: DOCUMENTS
# ═══════════════════════════════════════
def page_documents():
    r = get_role()
    st.markdown(
        "<div class='pro-header'><h1>📁 Document Center</h1>"
        "<p>Project files and deliverables</p></div>",
        unsafe_allow_html=True,
    )
    if not SUPABASE_OK:
        st.warning("Document storage requires Supabase.")
        return

    perm = get_permission("Documents", r)
    if perm in ("upload", "full"):
        with st.expander("📤 Upload Document"):
            cat = st.selectbox("Category", ["Proposal Draft", "Meeting Minutes", "Budget", "Research", "Deliverable", "Template", "Other"])
            desc = st.text_input("Description")
            uf = st.file_uploader("File", type=["pdf", "docx", "xlsx", "pptx", "png", "jpg", "csv", "md", "txt"])
            if uf and st.button("📤 Upload", type="primary"):
                sp = f"documents/{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uf.name}"
                if upload_to_storage(uf.getvalue(), sp, uf.type or "application/octet-stream"):
                    save_document_metadata(uf.name, uf.type, uf.size, cat, desc, get_name(), get_country(), sp)
                    st.success(f"✅ Uploaded: {uf.name}")
                    st.rerun()

    try:
        docs = sb().table("documents").select("*").eq("is_active", True).order("created_at", desc=True).execute().data or []
        if docs:
            st.subheader(f"📄 Documents ({len(docs)})")
            for d in docs:
                with st.expander(f"📄 {d.get('file_name','')} — {d.get('category','')}"):
                    st.markdown(f"**Category:** {d.get('category','')} | **By:** {d.get('uploaded_by','')} | **Date:** {str(d.get('created_at',''))[:10]}")
                    if d.get("description"):
                        st.markdown(f"*{d['description']}*")
                    dc1, dc2 = st.columns(2)
                    with dc1:
                        data = download_from_storage(d.get("storage_path", ""))
                        if data:
                            st.download_button("📥 Download", data, d.get("file_name", "file"), use_container_width=True)
                    if r == "Admin":
                        with dc2:
                            if st.button("🗑️ Delete", key=f"del_doc_{d.get('id')}"):
                                delete_from_storage(d.get("storage_path", ""))
                                sb().table("documents").update({"is_active": False}).eq("id", d["id"]).execute()
                                st.success("Deleted!")
                                st.rerun()
        else:
            st.info("No documents yet.")
    except Exception as e:
        st.error(f"Error: {e}")


# ═══════════════════════════════════════
# PAGE: AI CENTER
# ═══════════════════════════════════════
def page_ai_center():
    st.markdown(
        "<div class='pro-header'><h1>🧠 AI Decision Center</h1>"
        "<p>AI-powered feedback analysis and proposal improvement</p></div>",
        unsafe_allow_html=True,
    )
    if not AI_ENABLED:
        st.warning("AI not configured. Add API key to secrets.")
        return

    tab1, tab2, tab3, tab4 = st.tabs(["🔍 Analyze", "✍️ Revise", "📊 AI Log", "📋 Improvements"])

    with tab1:
        fbl = db_get_partner_feedback()
        open_fb = [f for f in fbl if f.get("status") == "Open"]
        if not open_fb:
            st.info("No open feedback.")
            return
        sel = st.selectbox(
            "Select Feedback",
            [f"#{f.get('id')} [{f.get('section','')}] {f.get('feedback', f.get('content',''))[:80]}..." for f in open_fb],
        )
        if sel:
            idx = int(sel.split("#")[1].split(" ")[0])
            fb = next((f for f in open_fb if f.get("id") == idx), None)
            if fb and st.button("🧠 Analyze with AI", type="primary"):
                with st.spinner("AI analyzing..."):
                    sec = fb.get("section", "")
                    sd = _find_section_for_feedback(sec)
                    sc = sd.get("content", "") if sd else ""
                    result = ai_analyze_feedback_v2(fb.get("feedback", fb.get("content", "")), sec, sc)
                    st.json(result)
                    db_log_ai_decision(fb.get("id"), result.get("decision", ""), result.get("confidence", 0), result.get("reasoning", ""), result.get("target_section", ""))
                    if result.get("decision") == "integrate" and result.get("suggested_text"):
                        st.markdown("### 📝 Suggested Revision")
                        st.markdown(result["suggested_text"])
                        if st.button("✅ Apply to Proposal"):
                            sk = sd.get("section_key", sec) if sd else sec
                            nv = db_update_section_content(sk, result["suggested_text"], "AI Engine", fb.get("id"))
                            db_log_improvement(fb.get("id"), sec, sc, result["suggested_text"], result.get("reasoning", ""), "ai_integrate", "AI Engine")
                            db_update_feedback_status(fb.get("id"), "Accepted", f"AI integrated (v{nv})")
                            st.success(f"Applied v{nv}!")
                            st.rerun()

    with tab2:
        sections = db_get_proposal_sections()
        if not sections:
            st.info("No proposal sections loaded.")
            return
        sk_sel = st.selectbox("Section", [f"{s.get('section_key')} — {s.get('section_title','')}" for s in sections])
        if sk_sel:
            sk = sk_sel.split(" — ")[0]
            sd = db_get_section_by_key(sk)
            if sd:
                st.text_area("Current Content", sd.get("content", ""), height=150, disabled=True)
                fb_input = st.text_area("Feedback / Instruction", height=80)
                if fb_input and st.button("✍️ Generate", type="primary"):
                    with st.spinner("AI writing..."):
                        revised = ai_generate_section_revision(sk, sd.get("content", ""), fb_input)
                        st.markdown("### Revised Version")
                        st.markdown(revised)
                        if st.button("✅ Apply"):
                            nv = db_update_section_content(sk, revised, "AI Engine")
                            db_log_improvement(None, sd.get("section_title", ""), sd.get("content", ""), revised, fb_input, "ai_revision", "AI Engine")
                            st.success(f"Applied v{nv}!")
                            st.rerun()

    with tab3:
        decisions = db_get_ai_decisions()
        if decisions:
            st.dataframe(pd.DataFrame(decisions).head(50), use_container_width=True, hide_index=True)
        else:
            st.info("No AI decisions yet.")

    with tab4:
        imps = db_get_improvement_log()
        if imps:
            st.dataframe(pd.DataFrame(imps).head(50), use_container_width=True, hide_index=True)
        else:
            st.info("No improvements yet.")


# ═══════════════════════════════════════
# PAGE: ADMIN PANEL
# ═══════════════════════════════════════
def page_admin():
    st.markdown(
        "<div class='pro-header'><h1>🛡️ Admin Panel</h1>"
        "<p>System administration and proposal management</p></div>",
        unsafe_allow_html=True,
    )
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Overview", "📝 Sections", "📥 Import", "📤 Export"])

    with tab1:
        sc1, sc2, sc3, sc4 = st.columns(4)
        with sc1:
            st.metric("Database", "Connected" if SUPABASE_OK else "Local")
        with sc2:
            st.metric("AI Engine", "Active" if AI_ENABLED else "Inactive")
        with sc3:
            st.metric("Export", "Ready" if EXPORT_OK else "N/A")
        with sc4:
            st.metric("Role", get_role())
        dc1, dc2, dc3, dc4 = st.columns(4)
        with dc1:
            st.metric("Partner FB", len(db_get_partner_feedback()))
        with dc2:
            st.metric("Patient FB", len(db_get_patient_feedback()))
        with dc3:
            st.metric("AI Decisions", len(db_get_ai_decisions()))
        with dc4:
            st.metric("Meetings", len(db_get_meetings()))

    with tab2:
        sections = db_get_proposal_sections()
        if sections:
            for s in sections:
                with st.expander(f"📄 {s.get('section_title','')} (v{s.get('version',1)})"):
                    new_c = st.text_area("Content", s.get("content", ""), height=200, key=f"sec_{s.get('section_key')}")
                    if st.button("💾 Save", key=f"save_{s.get('section_key')}"):
                        nv = db_update_section_content(s["section_key"], new_c, get_name())
                        st.success(f"Saved v{nv}!")
                        st.rerun()
        else:
            st.info("No sections. Import a proposal below.")

    with tab3:
        md_text = st.text_area("Paste Markdown", height=300, placeholder="## Project Summary\n...")
        if md_text and st.button("📥 Parse & Import", type="primary"):
            parsed = parse_proposal_md(md_text)
            if parsed:
                for sk, content in parsed.items():
                    st.markdown(f"**{sk}**: {len(content)} chars")
                if st.button("✅ Confirm Import"):
                    for sk, content in parsed.items():
                        title = sk.replace("_", " ").title()
                        if SUPABASE_OK:
                            try:
                                sb().table("proposal_sections").upsert({
                                    "section_key": sk, "section_title": title,
                                    "content": content, "version": 1,
                                    "last_updated_by": get_name(), "is_active": True,
                                }).execute()
                            except Exception:
                                pass
                    st.success(f"Imported {len(parsed)} sections!")
                    st.rerun()
            else:
                st.warning("Could not parse. Use ## headings.")

    with tab4:
        if EXPORT_OK:
            ec1, ec2 = st.columns(2)
            with ec1:
                xl = generate_feedback_excel()
                if xl:
                    st.download_button(
                        "📥 Full Excel",
                        xl,
                        f"oncoconnect_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            with ec2:
                pdf = generate_feedback_pdf()
                if pdf:
                    st.download_button(
                        "📥 Full PDF",
                        pdf,
                        f"oncoconnect_{datetime.now().strftime('%Y%m%d')}.pdf",
                        "application/pdf",
                        use_container_width=True,
                    )
            st.markdown("---")
            meetings = db_get_meetings()
            if meetings:
                mdf = pd.DataFrame(meetings)
                csv = mdf.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Meetings CSV", csv, "meetings.csv", "text/csv", use_container_width=True)
        else:
            st.warning("Install openpyxl and fpdf2 for export.")


# ═══════════════════════════════════════
# PAGE: USER MANAGEMENT
# ═══════════════════════════════════════
def page_user_management():
    st.markdown(
        "<div class='pro-header'><h1>👥 User Management</h1>"
        "<p>Manage platform users and permissions</p></div>",
        unsafe_allow_html=True,
    )
    if not SUPABASE_OK:
        st.info("Showing built-in users.")
        df = pd.DataFrame([
            {"Username": k, "Name": v["name"], "Role": v["role"], "Country": v["country"], "Org": v["org"]}
            for k, v in USERS_DB.items()
        ])
        st.dataframe(df, use_container_width=True, hide_index=True)
        return

    try:
        users = sb().table("app_users").select("*").execute().data or []
        if users:
            df = pd.DataFrame(users)
            show_cols = [c for c in ["username", "display_name", "role", "country", "organisation", "is_active", "last_login"] if c in df.columns]
            st.dataframe(df[show_cols], use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("### ➕ Add User")
        with st.form("add_user"):
            uc1, uc2 = st.columns(2)
            with uc1:
                nu = st.text_input("Username")
                nn = st.text_input("Display Name")
                np2 = st.text_input("Password", type="password")
            with uc2:
                nr = st.selectbox("Role", ["Admin", "Partner", "Patient"])
                nc = st.selectbox("Country", ["Turkey", "Poland", "Spain", "All", "N/A"])
                no = st.text_input("Organisation")
            if st.form_submit_button("➕ Add User", type="primary"):
                if nu and nn and np2:
                    try:
                        sb().table("app_users").insert({
                            "username": nu.lower(), "password_hash": np2,
                            "display_name": nn, "role": nr, "country": nc,
                            "organisation": no, "is_active": True,
                        }).execute()
                        st.success(f"User '{nu}' created!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")
    except Exception as e:
        st.error(f"Error: {e}")


# ═══════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════
def render_sidebar():
    with st.sidebar:
        st.markdown(
            "<div style='text-align:center;padding:1rem 0;'>"
            "<div style='font-size:2.5rem;'>🧬</div>"
            "<h2 style='color:#2ABFBF;margin:0.3rem 0;font-size:1.3rem;'>OncoConnect</h2>"
            "<p style='color:#64748b;font-size:0.75rem;margin:0;'>Co-Creation Hub v4.0</p></div>",
            unsafe_allow_html=True,
        )
        st.markdown("---")

        r = get_role()
        badge = ROLE_BADGES.get(r, r)
        cn = get_country()
        flag = FLAGS.get(cn, "🌍")

        st.markdown(
            f"<div style='background:rgba(42,191,191,0.1);border-radius:12px;padding:1rem;margin-bottom:1rem;'>"
            f"<p style='margin:0;color:#2ABFBF;font-weight:600;'>{get_name()}</p>"
            f"<p style='margin:0.2rem 0;color:#94a3b8;font-size:0.85rem;'>{badge}</p>"
            f"<p style='margin:0;color:#94a3b8;font-size:0.85rem;'>{flag} {cn} | {get_org()}</p></div>",
            unsafe_allow_html=True,
        )

        all_pages = [
            "Dashboard", "Work Packages", "Gantt Chart", "Partners",
            "Partner Feedback", "Patient Feedback", "Approval Status",
            "Announcements", "Documents", "Meetings",
            "🧠 AI Center", "Admin Panel", "User Management",
        ]
        pages = [p for p in all_pages if check_access(p, r)]

        page_icons = {
            "Dashboard": "📊", "Work Packages": "📦", "Gantt Chart": "📅",
            "Partners": "🤝", "Partner Feedback": "💬", "Patient Feedback": "💚",
            "Approval Status": "🗳️", "Announcements": "📢", "Documents": "📁",
            "Meetings": "📅", "🧠 AI Center": "🧠", "Admin Panel": "🛡️",
            "User Management": "👥",
        }

        page = st.radio(
            "Navigation", pages,
            format_func=lambda x: f"{page_icons.get(x, '')} {x}",
            label_visibility="collapsed",
        )

        st.markdown("---")

        ap = db_get_approvals()
        an = sum(1 for v in ap.values() if v)
        st.markdown(
            f"<div style='font-size:0.8rem;color:#94a3b8;'>"
            f"<p>🔗 DB: {'✅' if SUPABASE_OK else '⚠️'} | 🧠 AI: {'✅' if AI_ENABLED else '⚠️'}</p>"
            f"<p>🗳️ Approvals: {an}/3 | 📤 Export: {'✅' if EXPORT_OK else '⚠️'}</p></div>",
            unsafe_allow_html=True,
        )

        # Next meeting in sidebar
        meetings = db_get_meetings()
        today_str = str(date.today())
        upcoming = sorted(
            [m for m in meetings if str(m.get("meeting_date", "")) >= today_str],
            key=lambda x: str(x.get("meeting_date", "")),
        )
        if upcoming:
            nm = upcoming[0]
            st.markdown(
                f"<div style='background:rgba(45,140,255,0.1);border-radius:10px;padding:0.8rem;margin-top:0.5rem;'>"
                f"<p style='margin:0;font-size:0.7rem;color:#64748b;text-transform:uppercase;letter-spacing:1px;'>Next Meeting</p>"
                f"<p style='margin:0.2rem 0;color:#2D8CFF;font-weight:600;font-size:0.85rem;'>{nm.get('title','')[:30]}</p>"
                f"<p style='margin:0;color:#94a3b8;font-size:0.8rem;'>📅 {nm.get('meeting_date','')} 🕐 {str(nm.get('start_time',''))[:5]}</p></div>",
                unsafe_allow_html=True,
            )

        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            logout()

        st.markdown(
            "<p style='text-align:center;font-size:0.7rem;color:#475569;margin-top:1rem;'>"
            "© 2025 OncoConnect<br>Erasmus+ KA210</p>",
            unsafe_allow_html=True,
        )

        return page


# ═══════════════════════════════════════
# MAIN
# ═══════════════════════════════════════
def main():
    init_session()
    if not render_login():
        return
    inject_pro_css()

    page = render_sidebar()
    wp_df, pf = load_static()

    if page == "Dashboard":
        page_dashboard(wp_df, pf)
    elif page == "Work Packages":
        page_work_packages(wp_df)
    elif page == "Gantt Chart":
        page_gantt(wp_df)
    elif page == "Partners":
        page_partners(pf)
    elif page == "Partner Feedback":
        page_partner_feedback()
    elif page == "Patient Feedback":
        page_patient_feedback()
    elif page == "Approval Status":
        page_approval()
    elif page == "Announcements":
        page_announcements()
    elif page == "Documents":
        page_documents()
    elif page == "Meetings":
        page_meetings()
    elif page == "🧠 AI Center":
        page_ai_center()
    elif page == "Admin Panel":
        page_admin()
    elif page == "User Management":
        page_user_management()


if __name__ == "__main__":
    main()
